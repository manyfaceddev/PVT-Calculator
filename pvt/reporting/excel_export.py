"""
pvt/reporting/excel_export.py — Write ReportTable sections to a single-sheet
ADRIC-styled Excel workbook (openpyxl).

Layout, top to bottom: a navy/white-bold title banner, a "Sample
Information" block built from the `Sample`, then each `ReportTable` in turn
as a bold section-title row followed by its rows as three columns
(label/value/unit). Any row whose value is a QC severity string
("PASS"/"REVIEW"/"FAIL") gets its value cell filled green/amber/red.
"""

from pathlib import Path
from typing import IO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pvt.core.sample import Sample
from pvt.reporting.tables import ReportTable

_NAVY = "00205B"
_WHITE = "FFFFFF"

_SEVERITY_FILL: dict[str, str] = {
    "PASS": "38A169",
    "REVIEW": "DD9A0A",
    "FAIL": "E53E3E",
}

_HEADER_FONT = Font(color=_WHITE, bold=True, size=14)
_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_SECTION_FONT = Font(bold=True)

_LAST_COLUMN = 3
"""Number of columns a banner/section-title row is merged across, and that a
data row's label/value/unit occupy (A/B/C)."""


def _fmt_depth(depth_ft_md: float | None) -> str:
    """Format a sample's MD depth, or "N/A" when not recorded."""
    return f"{depth_ft_md:.1f}" if depth_ft_md is not None else "N/A"


def _write_banner(ws: Worksheet, row: int, text: str) -> int:
    """Write the navy/white-bold title banner, merged across all columns.

    Returns the next free row.
    """
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_LAST_COLUMN)
    return row + 1


def _write_section_title(ws: Worksheet, row: int, title: str) -> int:
    """Write a bold section-title row, merged across all columns.

    Returns the next free row.
    """
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_LAST_COLUMN)
    return row + 1


def _write_row(ws: Worksheet, row: int, label: str, value: str, unit: str) -> int:
    """Write a three-column (label/value/unit) row.

    The value cell is filled green/amber/red when `value` is a QC severity
    string ("PASS"/"REVIEW"/"FAIL"). Returns the next free row.
    """
    ws.cell(row=row, column=1, value=label)
    value_cell = ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=3, value=unit)
    color = _SEVERITY_FILL.get(value)
    if color is not None:
        value_cell.fill = PatternFill("solid", fgColor=color)
    return row + 1


def _write_sample_block(ws: Worksheet, sample: Sample, row: int) -> int:
    """Write the "Sample Information" section for `sample`.

    Returns the next free row.
    """
    row = _write_section_title(ws, row, "Sample Information")
    fields = [
        ("Sample ID", sample.sample_id, ""),
        ("Well", sample.well, ""),
        ("Field", sample.field_name, ""),
        ("Reservoir", sample.reservoir, ""),
        ("Depth (MD)", _fmt_depth(sample.depth_ft_md), "ft"),
        ("Fluid Type", sample.fluid_type, ""),
        ("Cylinder", sample.cylinder, ""),
        ("Client", sample.client, ""),
        ("Project", sample.project, ""),
    ]
    for label, value, unit in fields:
        row = _write_row(ws, row, label, value, unit)
    return row


def write_report(
    path: str | Path | IO[bytes], tables: list[ReportTable], *, title: str, sample: Sample
) -> None:
    """Write `tables` (plus a Sample Information block) to a single-sheet
    ADRIC-styled Excel workbook at `path`.

    Args:
        path: Output .xlsx path, or a writable binary file-like (e.g. an
            in-memory `io.BytesIO`, as `ui.common.components.report_download`
            uses to build a download without touching disk) -- `Workbook.
            save` (openpyxl) accepts either transparently; this signature
            just documents that.
        tables: Report sections to write, in order (e.g. from
            `pvt.reporting.tables.flash_tables`/`recombination_tables`).
        title: Report title, shown in the navy header banner.
        sample: Sample metadata, written as a "Sample Information" section.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    row = _write_banner(ws, 1, title)
    row += 1  # spacer
    row = _write_sample_block(ws, sample, row)
    row += 1  # spacer

    for table in tables:
        row = _write_section_title(ws, row, table.title)
        for report_row in table.rows:
            row = _write_row(ws, row, report_row.label, report_row.value, report_row.unit)
        row += 1  # spacer between tables

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    wb.save(path)
