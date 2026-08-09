"""Excel importer for `ADRIC_LiveOil_Preparation_Calc_v4.1.xlsx`.

Reads all five sheets of the filled workbook — `Sample_Info`,
`STO_Composition`, `Gas_Composition`, `Recombination`, and `Loading_Volumes`
— and produces a `LiveOilImport` ready to hand to
`pvt.experiments.recombination.molar.molar_split`/`wellstream` and
`pvt.experiments.recombination.loading.plan_loading`.

Dissection facts (confirmed by direct cell dump of
`tests/fixtures/workbooks/ADRIC_LiveOil_Preparation_Calc_v4.1.xlsx`):

BLOCK A - RECOMBINATION INPUTS (`Recombination!B5:B12`):

    | Cell | Label                          | -> field    |
    |------|----------------------------------|-------------|
    | B5   | GOR (scf/STB)                    | gor         |
    | B6   | GOR Type ("Separator"/"Stock Tank") | gor_basis |
    | B7   | Shrinkage Factor B_st             | shrinkage   |
    | B12  | Z at Standard Conditions           | z_std       |

    B6 dropdown text -> `GorBasis`: mapped verbatim ("Separator" ->
    `GorBasis.SEPARATOR`, "Stock Tank" -> `GorBasis.STOCK_TANK`) rather than
    swapped, per the conventional direction documented at
    `pvt.experiments.recombination.molar.molar_split` (D-018: the workbook's
    own B8 formula divides on the *opposite* branch — see
    `docs/excel-deviations.md` D-018, status NEEDS SWEJ RULING). This
    fixture's B7 (shrinkage) = 1.0, so the two directions are numerically
    identical here and the golden numbers are unaffected either way.

BLOCK B - STO PROPERTIES (`STO_Composition!B5`, `D65`):

    | Cell | Label                        | -> field         |
    |------|-------------------------------|-------------------|
    | B5   | STO Density @ 60°F reported   | sto_density_60f   |
    | D65  | C36+ MW (editable, row 65)    | c36_mw            |

    `c36_mw` overrides `KATZ_FIROOZABADI`'s default C36+ MW (636.4, see
    D-001) via `library.with_c36_mw(c36_mw)` before either composition
    stream is built — both `STO_Composition` and `Gas_Composition` share
    that per-sample library instance. Confirmed necessary: building the
    streams against the unmodified canonical library reproduces the brief's
    `f_gas` golden only to ~1e-4 (borderline); overriding C36+ MW first
    tightens that to ~1e-6 and exactly matches `tests/fixtures/sa372.py`'s
    `STO_MW_FROM_MOL`/`STO_C36_MW`, the fixture already proven against this
    same workbook by `tests/golden/test_molar_recombination_sa372.py`.

BLOCK C - COMPOSITIONS (`STO_Composition!` and `Gas_Composition!`, row 14
header, rows 15-65, 51 components; col B = component code, col I = Mol%
(INPUT) — the only composition column this importer reads; col J (Wt%
INPUT) is not consumed by any downstream calc.py in this phase):

    Naming differs slightly from the Flash v6.1 template (which needed 3
    aliases) — this template needs 4, discovered by diffing the sheet's
    column-B text against `pvt.core.components.KATZ_FIROOZABADI.codes`:

        "Neo-C5"     -> "NeoC5"      (workbook Component col: "Neo-Pentane")
        "Cyclohex"   -> "CycloC6"    (workbook Component col: "Cyclohexane")
        "E-Benzene"  -> "EBenzene"   (workbook Component col: "E-Benzene")
        "M/P-Xylene" -> "MP-Xylene"  (workbook Component col: "M/P-Xylene")

    ("O-Xylene" needs no alias here — the workbook already spells it exactly
    like the library code, unlike Flash's "OXylene" -> "O-Xylene".)

    This template's 51-row composition block also omits "TMB124"
    (1,2,4-Trimethylbenzene) entirely — KATZ_FIROOZABADI carries 52 codes,
    this sheet only 51. That component is simply absent from every row read
    here rather than aliased to something else; `CompositionStream` doesn't
    require every library code to be present, so the resulting stream is
    just silently missing that one component, matching what the lab
    actually reported. Cross-checked exactly (dict equality, both streams)
    against `tests/fixtures/sa372.py`'s `STO_MOL_PCT`/`GAS_MOL_PCT`.

BLOCK D - LOADING (`Loading_Volumes!B5:B12`):

    | Cell | -> LoadingInputs field       |
    |------|------------------------------|
    | B5   | cylinder_volume_cc           |
    | B6   | target_oil_cc                |
    | B7   | oil_load_p_psig              |
    | B8   | oil_load_t_f                 |
    | B9   | gas_load_p_psig              |
    | B10  | gas_load_t_f                 |
    | B11  | z_gas_load                   |
    | B12  | sto_density_at_load_g_cc     |

    (B13, "BSW / Water content of loaded oil", has no `LoadingInputs` field
    and is not read.)

BLOCK E - SAMPLE METADATA (`Sample_Info!B5:B9`/`E5:E9`):

    | Cell | Label                   | -> Sample field | Cell | Label                | -> Sample field |
    |------|--------------------------|------------------|------|------------------------|------------------|
    | B5   | Company / Client         | client           | E5   | Field                  | field_name       |
    | B6   | Well                     | well             | E6   | Reservoir              | reservoir        |
    | B7   | Sample ID                | sample_id        | E7   | Chamber / Cylinder No. | cylinder         |
    | B8   | Sampling Depth (MD ft)   | depth_ft_md      | E8   | Sampling Date          | *unused*         |
    | B9   | Fluid Type               | fluid_type       | E9   | Project No.            | project          |

    Unlike the Flash v6.1 importer, this template carries a `Reservoir`
    field (`E6`) directly — `Sample.reservoir` is populated here rather than
    left `""`.

Wrong-file / shifted-layout detection: any of the 5 required sheets missing,
or a title/header-row anchor cell that doesn't match the expected text,
raises `InputValidationError`. The Flash v6.1 workbook is rejected by the
missing-sheet check alone (it has no `Sample_Info`/`STO_Composition`/
`Gas_Composition`/`Loading_Volumes` sheet) — but note it *does* carry a sheet
literally named "Recombination" with an entirely different layout, which is
why the title-anchor check on that sheet is kept too rather than relying on
sheet-name presence alone.

Import-boundary guard: any composition cell in rows 15-65 (col I) that reads
as negative raises `InputValidationError`, same rationale as the Flash v6.1
importer (Task 7) — `CompositionStream` itself does not validate value sign.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from pvt.core.components import KATZ_FIROOZABADI, ComponentLibrary
from pvt.core.composition import CompositionStream
from pvt.core.exceptions import InputValidationError
from pvt.core.sample import Sample
from pvt.experiments.recombination.loading import LoadingInputs
from pvt.experiments.recombination.molar import GorBasis

# Workbook code -> pvt.core.components.KATZ_FIROOZABADI code. See module
# docstring BLOCK C for the discovery process (diff against KF.codes).
_ALIAS: dict[str, str] = {
    "Neo-C5": "NeoC5",
    "Cyclohex": "CycloC6",
    "E-Benzene": "EBenzene",
    "M/P-Xylene": "MP-Xylene",
}

_COMPOSITION_FIRST_ROW = 15
_COMPOSITION_LAST_ROW = 65

_REQUIRED_SHEETS = (
    "Sample_Info",
    "STO_Composition",
    "Gas_Composition",
    "Recombination",
    "Loading_Volumes",
)

# Anchor cells checked for wrong-file / shifted-header detection, keyed by
# (sheet, address).
_EXPECTED_LABELS: dict[tuple[str, str], str] = {
    ("Sample_Info", "A1"): (
        "ADRIC — LIVE OIL PREPARATION CALCULATION (v4.1, MG-0180 methodology, "
        "live formulas + physics notes)"
    ),
    ("STO_Composition", "A1"): (
        "STOCK TANK OIL — COMPOSITION & PROPERTIES (to C36+, Katz-Firoozabadi reference)"
    ),
    ("STO_Composition", "B14"): "Code",
    ("STO_Composition", "I14"): "Mol% (INPUT)",
    ("Gas_Composition", "A1"): (
        "SEPARATOR GAS — COMPOSITION & PROPERTIES (to C36+, Katz-Firoozabadi reference)"
    ),
    ("Gas_Composition", "B14"): "Code",
    ("Gas_Composition", "I14"): "Mol% (INPUT)",
    ("Recombination", "A1"): "RECOMBINATION — WELLSTREAM COMPOSITION & K-VALUE VALIDATION",
    ("Loading_Volumes", "A1"): "LOADING VOLUMES — CYLINDER CHARGING FOR LIVE OIL PREPARATION",
}

# D-018: Recombination!B6 dropdown text -> GorBasis, mapped verbatim pending
# Swej's ruling on the basis-conversion direction (docs/excel-deviations.md
# D-018). This fixture's B7 (shrinkage) = 1.0, so goldens are invariant
# either way.
_BASIS_MAP: dict[str, GorBasis] = {
    "Separator": GorBasis.SEPARATOR,
    "Stock Tank": GorBasis.STOCK_TANK,
}


@dataclass(frozen=True)
class LiveOilImport:
    """Everything read from a filled LiveOil v4.1 workbook."""

    gor: float
    gor_basis: GorBasis
    shrinkage: float
    z_std: float
    sto_density_60f: float
    c36_mw: float
    sto_stream: CompositionStream
    gas_stream: CompositionStream
    loading: LoadingInputs
    sample: Sample


def read(path: str | Path) -> LiveOilImport:
    """Import a filled `ADRIC_LiveOil_Preparation_Calc_v4.1.xlsx` workbook.

    Args:
        path: Path to the filled workbook.

    Returns:
        LiveOilImport with the recombination GOR inputs, both composition
        streams (STO and gas, mol% basis, C36+ MW overridden per D65),
        loading-cylinder inputs, and sample metadata.

    Raises:
        InputValidationError: the workbook is missing one of the 5 required
            sheets, a title/header anchor cell doesn't match the expected
            v4.1 template text, the `Recombination!B6` GOR-basis text isn't
            recognized, or a composition cell in rows 15-65 (col I) reads as
            negative.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        missing = [sheet for sheet in _REQUIRED_SHEETS if sheet not in wb.sheetnames]
        if missing:
            raise InputValidationError(
                [
                    f"{path}: missing required sheet(s) {missing} — not an "
                    "ADRIC LiveOil Preparation Calc v4.1 workbook"
                ]
            )
        _check_layout(wb)

        c36_mw = _num(wb["STO_Composition"], "D65")
        library = KATZ_FIROOZABADI.with_c36_mw(c36_mw)

        gor, gor_basis, shrinkage, z_std = _read_recombination_inputs(wb["Recombination"])
        sto_density_60f = _num(wb["STO_Composition"], "B5")
        sto_stream = _read_composition(wb["STO_Composition"], library)
        gas_stream = _read_composition(wb["Gas_Composition"], library)
        loading = _read_loading(wb["Loading_Volumes"])
        sample = _read_sample(wb["Sample_Info"])
    finally:
        wb.close()

    return LiveOilImport(
        gor=gor,
        gor_basis=gor_basis,
        shrinkage=shrinkage,
        z_std=z_std,
        sto_density_60f=sto_density_60f,
        c36_mw=c36_mw,
        sto_stream=sto_stream,
        gas_stream=gas_stream,
        loading=loading,
        sample=sample,
    )


def _check_layout(wb: Any) -> None:
    """Raise InputValidationError if any anchor cell's text has shifted."""
    errors = [
        f"{sheet}!{addr}: expected {expected!r}, found {wb[sheet][addr].value!r}"
        for (sheet, addr), expected in _EXPECTED_LABELS.items()
        if wb[sheet][addr].value != expected
    ]
    if errors:
        raise InputValidationError(errors)


def _num(ws: Any, addr: str) -> float:
    return float(ws[addr].value)


def _read_recombination_inputs(ws: Any) -> tuple[float, GorBasis, float, float]:
    gor = _num(ws, "B5")
    basis_text = str(ws["B6"].value)
    if basis_text not in _BASIS_MAP:
        raise InputValidationError(
            [f"Recombination!B6: unrecognized GOR basis {basis_text!r}"]
        )
    gor_basis = _BASIS_MAP[basis_text]
    shrinkage = _num(ws, "B7")
    z_std = _num(ws, "B12")
    return gor, gor_basis, shrinkage, z_std


def _read_loading(ws: Any) -> LoadingInputs:
    return LoadingInputs(
        cylinder_volume_cc=_num(ws, "B5"),
        target_oil_cc=_num(ws, "B6"),
        oil_load_p_psig=_num(ws, "B7"),
        oil_load_t_f=_num(ws, "B8"),
        gas_load_p_psig=_num(ws, "B9"),
        gas_load_t_f=_num(ws, "B10"),
        z_gas_load=_num(ws, "B11"),
        sto_density_at_load_g_cc=_num(ws, "B12"),
    )


def _read_sample(ws: Any) -> Sample:
    return Sample(
        sample_id=str(ws["B7"].value),
        well=str(ws["B6"].value),
        field_name=str(ws["E5"].value),
        reservoir=str(ws["E6"].value),
        depth_ft_md=_num(ws, "B8"),
        fluid_type=str(ws["B9"].value),
        cylinder=str(ws["E7"].value),
        client=str(ws["B5"].value),
        project=str(ws["E9"].value),
    )


def _read_composition(ws: Any, library: ComponentLibrary) -> CompositionStream:
    mol: dict[str, float] = {}
    errors: list[str] = []

    for row in range(_COMPOSITION_FIRST_ROW, _COMPOSITION_LAST_ROW + 1):
        raw_code = ws[f"B{row}"].value
        code = _ALIAS.get(raw_code, raw_code)
        value = ws[f"I{row}"].value

        if value < 0:
            errors.append(f"{ws.title}!row {row} ({code}): negative Mol% value {value}")

        if value:
            mol[code] = value

    if errors:
        raise InputValidationError(errors)

    return CompositionStream(library=library, mol_pct=mol)
