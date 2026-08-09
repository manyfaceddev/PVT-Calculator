"""Excel importer for `ADRIC_Flash_Separation_Calc_v6.1.xlsx`.

Reads the filled workbook's `Volumetrics_Master` sheet — the only sheet that
carries yellow (input) cells in this template; `Component_Properties`,
`Recombination`, and `Plus_Properties_Report` are all downstream/computed —
and produces a `FlashImport` ready to hand to
`pvt.experiments.flash.calc.calculate`.

Dissection facts (confirmed by direct cell dump of
`tests/fixtures/workbooks/ADRIC_Flash_Separation_Calc_v6.1.xlsx`,
`Volumetrics_Master`, rows 1-93):

BLOCK A - PROJECT & SAMPLE INFORMATION (row 5-8; consumed cells only —
metadata with no `Sample`/`FlashVolumetrics` field, e.g. reservoir P/T,
saturation pressure, analyst, sampling date, are not read here; they have no
home until `Study`/`CrossRef` are wired in a later phase):

    | Cell | Label                  | -> Sample field |
    |------|-------------------------|-----------------|
    | B5   | Client                  | client          |
    | B6   | Well / Field            | well, field_name (split on " / ") |
    | B7   | Sampling Depth (MD ft)  | depth_ft_md     |
    | H5   | Sample ID               | sample_id       |
    | H6   | Chamber / Cylinder      | cylinder        |
    | H8   | Project No.             | project         |
    | E8   | Fluid Type              | fluid_type      |

`Sample.reservoir` has no source cell in this block (only reservoir
*temperature*/*pressure*, which are conditions, not a reservoir name) and is
set to `""`.

BLOCK B - VOLUMETRIC MEASUREMENTS (row 11-21; yellow inputs only — B13, B16,
B19, and all of BLOCK C (B25-B37) are computed and are deliberately not
read):

    | Cell | Label                              | -> FlashVolumetrics field |
    |------|-------------------------------------|----------------------------|
    | B11  | Pump constant                       | pump_constant             |
    | E11  | Volume correction factor (VCF)      | vcf                        |
    | B12  | Initial pump reading (cc)           | pump_initial_cc            |
    | E12  | Final pump reading (cc)             | pump_final_cc              |
    | B14  | Stock tank oil volume V_sto (cc)    | v_sto_cc                   |
    | B15  | Oil tare weight (g)                 | oil_tare_g                 |
    | E15  | Final oil + tare weight (g)         | oil_gross_g                |
    | B17  | Gasometer factor                    | gasometer_factor           |
    | B18  | Initial gasometer reading (cc)      | gasometer_initial_cc       |
    | E18  | Final gasometer reading (cc)        | gasometer_final_cc         |
    | B20  | Gas temperature (C)                 | gas_temp_c                 |
    | B21  | Measured gas abs. pressure (mbar)   | gas_abs_pressure_mbar      |
    | E21  | Gas gravity (Air=1)                 | gas_gravity                |
    | E17  | Barometric pressure (mbar)          | *unused by engine* (ledger D-017: feeds the workbook's B25 `P_base` composite, which `calculate()` never consumes) |
    | E20  | Back pressure (mbar)                | *unused by engine* (ledger D-017, same as E17) |

BLOCK D - GC COMPOSITIONS (row 40 header, rows 41-92, 52 components):

    col B = component code, E = Gas Mol% (INPUT), F = Gas Wt% (INPUT),
    G = Oil Mol% (INPUT), H = Oil Wt% (INPUT). Row order matches
    `pvt.core.components.KATZ_FIROOZABADI.codes` exactly except for 3 codes
    that need an explicit alias (workbook code -> library code), the same
    map Task 3's `tests/fixtures/sa372_flash.py` generator proved against
    this same workbook:

        "Cyclohex"   -> "CycloC6"    (workbook Component col: "Cyclohexane")
        "MPXylenes"  -> "MP-Xylene"  (workbook Component col: "M/P-Xylenes")
        "OXylene"    -> "O-Xylene"   (workbook Component col: "O-Xylene")

Wrong-file / shifted-layout detection: a missing `Volumetrics_Master` sheet,
or a row-40/A1 header that doesn't match the expected text, raises
`InputValidationError` rather than silently reading garbage from the wrong
cells.

Import-boundary guard: any composition cell in rows 41-92 (cols E-H) that
reads as negative raises `InputValidationError` — the Excel-import boundary
is the right place to reject a malformed lab entry sheet, since
`CompositionStream` itself does not validate value sign (P0 final-review
carry-forward).
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from pvt.core.components import KATZ_FIROOZABADI
from pvt.core.composition import CompositionStream
from pvt.core.exceptions import InputValidationError
from pvt.core.sample import Sample
from pvt.experiments.flash.models import FlashVolumetrics

# Workbook code -> pvt.core.components.KATZ_FIROOZABADI code. Same 3-entry
# map proved by tests/fixtures/sa372_flash.py's generator script against
# this workbook (see that file's footer for the discovery process).
_ALIAS: dict[str, str] = {
    "Cyclohex": "CycloC6",
    "MPXylenes": "MP-Xylene",
    "OXylene": "O-Xylene",
}

_COMPOSITION_FIRST_ROW = 41
_COMPOSITION_LAST_ROW = 92

# Anchor cells checked for wrong-file / shifted-header detection.
_EXPECTED_LABELS: dict[str, str] = {
    "A1": "ATMOSPHERIC FLASH SEPARATION - WATER PUMP METHOD (v6.1)",
    "B40": "Code",
    "E40": "Gas Mol% (INPUT)",
    "F40": "Gas Wt% (INPUT)",
    "G40": "Oil Mol% (INPUT)",
    "H40": "Oil Wt% (INPUT)",
}


@dataclass(frozen=True)
class FlashImport:
    """Everything read from a filled Flash v6.1 workbook."""

    volumetrics: FlashVolumetrics
    oil_stream: CompositionStream
    gas_stream: CompositionStream
    sample: Sample


def read(path: str | Path) -> FlashImport:
    """Import a filled `ADRIC_Flash_Separation_Calc_v6.1.xlsx` workbook.

    Args:
        path: Path to the filled workbook.

    Returns:
        FlashImport with the volumetrics inputs, both GC composition
        streams (mol% and wt% bases), and the sample metadata.

    Raises:
        InputValidationError: the workbook is missing the
            `Volumetrics_Master` sheet, that sheet's header text doesn't
            match the expected v6.1 template layout, or a composition cell
            in rows 41-92 reads as negative.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if "Volumetrics_Master" not in wb.sheetnames:
            raise InputValidationError(
                [
                    f"{path}: missing required sheet 'Volumetrics_Master' — "
                    "not an ADRIC Flash Separation Calc v6.1 workbook"
                ]
            )
        ws = wb["Volumetrics_Master"]
        _check_layout(ws)
        volumetrics = _read_volumetrics(ws)
        sample = _read_sample(ws)
        gas_stream, oil_stream = _read_compositions(ws)
    finally:
        wb.close()

    return FlashImport(
        volumetrics=volumetrics, oil_stream=oil_stream, gas_stream=gas_stream, sample=sample
    )


def _check_layout(ws: Any) -> None:
    """Raise InputValidationError if the sheet's header text has shifted."""
    errors = [
        f"Volumetrics_Master!{addr}: expected {expected!r}, found {ws[addr].value!r}"
        for addr, expected in _EXPECTED_LABELS.items()
        if ws[addr].value != expected
    ]
    if errors:
        raise InputValidationError(errors)


def _num(ws: Any, addr: str) -> float:
    return float(ws[addr].value)


def _read_volumetrics(ws: Any) -> FlashVolumetrics:
    return FlashVolumetrics(
        pump_constant=_num(ws, "B11"),
        vcf=_num(ws, "E11"),
        pump_initial_cc=_num(ws, "B12"),
        pump_final_cc=_num(ws, "E12"),
        v_sto_cc=_num(ws, "B14"),
        oil_tare_g=_num(ws, "B15"),
        oil_gross_g=_num(ws, "E15"),
        gasometer_factor=_num(ws, "B17"),
        gasometer_initial_cc=_num(ws, "B18"),
        gasometer_final_cc=_num(ws, "E18"),
        gas_temp_c=_num(ws, "B20"),
        gas_abs_pressure_mbar=_num(ws, "B21"),
        gas_gravity=_num(ws, "E21"),
    )


def _read_sample(ws: Any) -> Sample:
    well, _sep, field_name = str(ws["B6"].value).partition(" / ")
    return Sample(
        sample_id=str(ws["H5"].value),
        well=well,
        field_name=field_name,
        reservoir="",
        depth_ft_md=_num(ws, "B7"),
        fluid_type=str(ws["E8"].value),
        cylinder=str(ws["H6"].value),
        client=str(ws["B5"].value),
        project=str(ws["H8"].value),
    )


def _read_compositions(ws: Any) -> tuple[CompositionStream, CompositionStream]:
    gas_mol: dict[str, float] = {}
    gas_wt: dict[str, float] = {}
    oil_mol: dict[str, float] = {}
    oil_wt: dict[str, float] = {}
    errors: list[str] = []

    for row in range(_COMPOSITION_FIRST_ROW, _COMPOSITION_LAST_ROW + 1):
        raw_code = ws[f"B{row}"].value
        code = _ALIAS.get(raw_code, raw_code)
        gm = ws[f"E{row}"].value
        gw = ws[f"F{row}"].value
        om = ws[f"G{row}"].value
        ow = ws[f"H{row}"].value

        for label, value in (
            ("Gas Mol%", gm), ("Gas Wt%", gw), ("Oil Mol%", om), ("Oil Wt%", ow)
        ):
            if value < 0:
                errors.append(
                    f"Volumetrics_Master!row {row} ({code}): negative {label} value {value}"
                )

        if gm:
            gas_mol[code] = gm
        if gw:
            gas_wt[code] = gw
        if om:
            oil_mol[code] = om
        if ow:
            oil_wt[code] = ow

    if errors:
        raise InputValidationError(errors)

    gas_stream = CompositionStream(library=KATZ_FIROOZABADI, mol_pct=gas_mol, wt_pct=gas_wt)
    oil_stream = CompositionStream(library=KATZ_FIROOZABADI, mol_pct=oil_mol, wt_pct=oil_wt)
    return gas_stream, oil_stream
