"""
tests/unit/experiments/test_cce_validate.py — CCE validation rule tests.

The happy-path fixture is built by reading the committed workbook
`tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx`
(sheet "CCE Calculation") directly with openpyxl -- there is no importer
yet (that is Task 5); this module's `_load_cce_happy_path` helper is a
test-only reader, not production code.

Cell-map notes (confirmed by loading the fixture with both
`data_only=True` and `data_only=False`):

- Stage table: rows 16-55 (40 stages). Columns:
    A = step (1..40)
    B = pressure, psig-as-entered (D3 pressure policy; see plan Global
        Constraints -- the engine treats this as absolute psia pass-through)
    C = ENTERED cell volume (cc)
    D = Relative Volume  (formula `=C{row}/$J$8`)      -- DERIVED
    E = Density (g/cm3)  (formula `=IF(...,$J$7/C{row},"")`) -- DERIVED
  Column determination: with `data_only=False`, C16 holds the raw float
  94.08822903 (no formula), while D16 and E16 hold formula strings that
  reference C16. So C is the entered volume; D/E are workbook-computed
  and are intentionally NOT read here (calc.py, Task 2, recomputes them).

- D5 = "Reservoir Pressure (psig)" = 3938.73    -> CceInputs.reservoir_p_psia
  D9 = "Visual Bubble Point (psig)" = 1155.73  -> CceInputs.psat_visual
  D10 = "Bubble Point Step #:" = 20             -> CceInputs.bubble_point_step

  D5 is a separate, independently-entered lab input from D7 "Working
  Pressure (psig)" = 7014.73 (row 16's pressure, the first stage) --
  added in Task 2 round 2 (controller adjudication, see
  pvt/experiments/cce/calc.py and docs/excel-deviations.md D-020) so
  that calc.py can anchor the reservoir->Psat mean compressibility the
  same way Mean Compressibility!H8 does.

  NOTE -- brief/plan discrepancy: the task-1 brief and the Phase 3a plan's
  Task 1/Task 5 cell maps both describe the visual Psat as cell **D8**.
  The fixture actually has D8 = "Measured HPHT Density (g/cm3)" = 0.72868
  and D9 = "Visual Bubble Point (psig)" = 1155.73 (rows 6-10 run
  Temperature, Working Pressure, Density, Visual Psat, Bubble Step). The
  *values* the brief quotes (visual Psat ~1155.73, bubble step 20) match
  D9/D10 exactly and unambiguously identify the correct cells, so this
  test (and pvt/experiments/cce/models.py's docstring) use D9. Flagged
  for the Task 5 importer author to use the corrected address. This is a
  plan-documentation offset, not a workbook defect, so it is not logged
  to docs/workbook-defect-review.md.

- Using the full 40-row table (not just rows 16-25) is required for the
  happy path to be self-consistent: D10 (bubble_point_step) = 20 must be
  a valid 1-based index into the stage table for the "bubble_point_step
  within stage range" rule to pass, which needs >= 20 stages present.
"""

import dataclasses
from pathlib import Path

import openpyxl

from pvt.experiments.cce.models import CceInputs, CceStage
from pvt.experiments.cce.validate import validate

WB = Path("tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx")


def _load_cce_happy_path() -> CceInputs:
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["CCE Calculation"]

    stages = []
    row = 16
    while ws[f"A{row}"].value is not None:
        stages.append(
            CceStage(
                step=int(ws[f"A{row}"].value),
                p=float(ws[f"B{row}"].value),
                v_cell_cc=float(ws[f"C{row}"].value),
            )
        )
        row += 1

    return CceInputs(
        t_res_f=float(ws["D6"].value),
        psat_visual=float(ws["D9"].value),
        bubble_point_step=int(ws["D10"].value),
        stages=tuple(stages),
        reservoir_p_psia=float(ws["D5"].value),
    )


HAPPY = _load_cce_happy_path()


def test_happy_path_from_fixture():
    assert validate(HAPPY) == []


def test_fixture_shape_sanity():
    # Sanity-checks on the fixture read itself, independent of validate().
    assert len(HAPPY.stages) == 40
    assert HAPPY.bubble_point_step == 20
    assert HAPPY.psat_visual == 1155.73
    assert HAPPY.stages[19].step == 20
    assert HAPPY.stages[19].p == 1155.73
    assert HAPPY.reservoir_p_psia == 3938.73


def test_too_few_stages_flagged():
    bad = dataclasses.replace(HAPPY, stages=HAPPY.stages[:1])
    errors = validate(bad)
    assert any("stage" in e.lower() for e in errors)


def test_pressure_not_descending_flagged():
    stages = list(HAPPY.stages)
    stages[5], stages[6] = stages[6], stages[5]  # break strict descent
    bad = dataclasses.replace(HAPPY, stages=tuple(stages))
    errors = validate(bad)
    assert any("descend" in e.lower() for e in errors)


def test_nonpositive_volume_flagged():
    stages = list(HAPPY.stages)
    stages[0] = dataclasses.replace(stages[0], v_cell_cc=0.0)
    bad = dataclasses.replace(HAPPY, stages=tuple(stages))
    errors = validate(bad)
    assert any("volume" in e.lower() for e in errors)


def test_bubble_point_step_out_of_range_flagged():
    bad = dataclasses.replace(HAPPY, bubble_point_step=41)
    errors = validate(bad)
    assert any("bubble_point_step" in e.lower() for e in errors)


def test_bubble_point_step_zero_flagged():
    bad = dataclasses.replace(HAPPY, bubble_point_step=0)
    errors = validate(bad)
    assert any("bubble_point_step" in e.lower() for e in errors)


def test_psat_consistency_advisory_flagged_but_nonblocking():
    bad = dataclasses.replace(HAPPY, psat_visual=HAPPY.psat_visual + 50.0)
    errors = validate(bad)
    assert any(e.startswith("consistency:") for e in errors)
    assert len(errors) == 1  # advisory only -- nothing else is broken


def test_psat_consistency_within_tolerance_not_flagged():
    bad = dataclasses.replace(HAPPY, psat_visual=HAPPY.psat_visual + 5.0)
    errors = validate(bad)
    assert not any(e.startswith("consistency:") for e in errors)


def test_psat_consistency_skipped_when_bubble_step_out_of_range():
    # Guard: must not raise (or double-report) when the picked row can't
    # be indexed at all.
    bad = dataclasses.replace(HAPPY, bubble_point_step=0, psat_visual=1.0)
    errors = validate(bad)
    assert not any(e.startswith("consistency:") for e in errors)


def test_t_res_f_too_low_flagged():
    bad = dataclasses.replace(HAPPY, t_res_f=-100.0)
    errors = validate(bad)
    assert any("t_res_f" in e.lower() for e in errors)


def test_t_res_f_too_high_flagged():
    bad = dataclasses.replace(HAPPY, t_res_f=600.0)
    errors = validate(bad)
    assert any("t_res_f" in e.lower() for e in errors)


def test_errors_accumulate():
    bad = dataclasses.replace(
        HAPPY,
        stages=HAPPY.stages[:1],  # too few stages
        t_res_f=1000.0,  # out of physical range
        bubble_point_step=99,  # out of range (also suppresses the consistency check)
        # neutralized: HAPPY.reservoir_p_psia (3938.73) would fall below
        # the single remaining stage's P (7014.73) once stages is
        # truncated above, tripping the new reservoir_p_psia advisory
        # rule as an unrelated 4th message -- keep this test's "3
        # independent violations" intent unambiguous.
        reservoir_p_psia=None,
    )
    errors = validate(bad)
    assert len(errors) == 3


def test_reservoir_p_psia_none_skips_rule():
    bad = dataclasses.replace(HAPPY, reservoir_p_psia=None)
    assert validate(bad) == []


def test_reservoir_p_psia_valid_not_flagged():
    # HAPPY.reservoir_p_psia = 3938.73, well within (last stage's P ..
    # RESERVOIR_P_MAX_PSIA) -- already implied by test_happy_path_from_fixture,
    # asserted explicitly here per the round-2 coverage requirement.
    errors = validate(HAPPY)
    assert not any("reservoir_p_psia" in e.lower() for e in errors)


def test_reservoir_p_psia_nonpositive_flagged_blocking():
    bad = dataclasses.replace(HAPPY, reservoir_p_psia=0.0)
    errors = validate(bad)
    reservoir_errors = [e for e in errors if "reservoir_p_psia" in e.lower()]
    assert len(reservoir_errors) == 1
    assert "must be > 0" in reservoir_errors[0]
    assert not reservoir_errors[0].startswith("consistency:")  # blocking, not advisory

    neg = dataclasses.replace(HAPPY, reservoir_p_psia=-500.0)
    errors_neg = validate(neg)
    assert any("reservoir_p_psia" in e.lower() for e in errors_neg)


def test_reservoir_p_psia_below_last_stage_flagged_advisory():
    # HAPPY's last stage (step 40) has P=218.0346 -- below that is
    # implausible (the reservoir pressure should be at or above every
    # pressure the expansion ever reaches).
    bad = dataclasses.replace(HAPPY, reservoir_p_psia=100.0)
    errors = validate(bad)
    assert any(
        e.startswith("consistency:") and "reservoir_p_psia" in e for e in errors
    )


def test_reservoir_p_psia_above_max_flagged_advisory():
    bad = dataclasses.replace(HAPPY, reservoir_p_psia=30_000.0)
    errors = validate(bad)
    assert any(
        e.startswith("consistency:") and "reservoir_p_psia" in e for e in errors
    )


def test_reservoir_p_psia_plausibility_check_skipped_when_no_stages():
    # Guard: a positive reservoir_p_psia with an empty stage table must
    # not crash trying to index stages[-1] -- the plausibility-band check
    # is simply skipped (the "too few stages" rule already flags the
    # empty table separately).
    bad = dataclasses.replace(HAPPY, stages=(), reservoir_p_psia=1000.0)
    errors = validate(bad)
    assert not any("reservoir_p_psia" in e.lower() for e in errors)
