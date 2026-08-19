"""
tests/unit/qc/test_cce_checks.py — Phase 3a Task 3: CCE fit QC +
resurrected checks (`pvt/qc/checks/polynomial_fit.py`,
`monotonic_compressibility.py`, `rho_v_constancy.py`).

Correctness tests use synthetic data constructed FROM a known polynomial
(exact reconstruction, rel=1e-9) or a hand-computed closed-form spread
(rho*V constancy has a trivial closed form, restated directly in the
test, same pattern as `test_cce_calc.py`'s
`test_mean_compressibility_helper_matches_two_point_form`). Grading-band
placement for the two polynomial-fit checks needs a real least-squares
solve to land precisely inside a narrow band (0.05-0.10% / 1.0-2.0%) --
those values are OFFLINE-COMPUTED by sweeping a perturbation parameter
through the real implementation during test authoring (see the comment
above each pair of band tests below) and embedding the confirmed result,
the same pattern Phase 2 used for
`test_hoffman_r2_review_band_three_component_fit` (see git commit
5c1e10f, "Task 6 review fix: cover hoffman_r2 REVIEW/FAIL bands, not just
perfect-fit PASS").

The final test in this module is integration-flavored: it runs all three
checks over the REAL fixture workbook's CCE engine output (built the same
way `tests/golden/test_cce_workbook.py` does) and asserts only that each
check returns a well-formed QCResult without raising -- per the task
brief, it does NOT assert specific severities, since the sheet's own data
may legitimately land in REVIEW on the resurrected/proposal-status checks.
Observed severities are reported in a comment there.
"""

import math
from pathlib import Path

import openpyxl
import pytest

from pvt.core.exceptions import InputValidationError
from pvt.experiments.cce.calc import calculate
from pvt.experiments.cce.models import CceInputs, CceStage
from pvt.qc.checks import monotonic_compressibility, polynomial_fit, rho_v_constancy
from pvt.qc.engine import QCResult, Severity, ThresholdRegistry

# --- registry / exports plumbing -------------------------------------------


def test_registry_has_new_cce_keys():
    reg = ThresholdRegistry()
    assert reg.get("cce_sp_fit_dev_pct") == (0.05, 0.10)
    assert reg.get("cce_tp_fit_dev_pct") == (1.0, 2.0)
    assert reg.get("cce_monotonic_violations") == (0.0, 1.0)
    assert reg.get("cce_rho_v_spread_pct") == (0.5, 1.0)


def test_checks_package_exports_new_modules():
    import pvt.qc.checks as checks

    assert "polynomial_fit" in checks.__all__
    assert "monotonic_compressibility" in checks.__all__
    assert "rho_v_constancy" in checks.__all__


# --- polynomial_fit: exact-reconstruction correctness -----------------------


def test_check_single_phase_reconstructs_exact_cubic_at_rel_1e9():
    # Points constructed FROM a known cubic in P -- an affine reparametrization
    # (centering/scaling P) preserves polynomial degree exactly, so a
    # degree-3 fit through exactly-cubic data must reconstruct it to
    # floating-point roundoff, regardless of the P range/units.
    a0, a1, a2, a3 = 1.05, -3.1e-5, 2.4e-9, -6.0e-14
    ps = [7000.0, 6500.0, 6000.0, 5500.0, 5000.0, 4500.0, 4000.0, 3500.0]
    points = [(p, a0 + a1 * p + a2 * p**2 + a3 * p**3) for p in ps]

    result = polynomial_fit.check_single_phase(points, degree=3)

    for (_, actual), fitted in zip(points, result.fitted):
        assert fitted == pytest.approx(actual, rel=1e-9)
    assert result.max_dev_pct == pytest.approx(0.0, abs=1e-6)
    assert result.qc.check_id == "cce_sp_fit_dev_pct"
    assert result.qc.severity == Severity.PASS
    assert len(result.coeffs) == 4


def test_check_two_phase_reconstructs_exact_quadratic_at_rel_1e9():
    b0, b1, b2 = 1.334, 8.41e-4, -2.21e-7
    ps = [1100.0, 950.0, 800.0, 650.0, 500.0, 350.0, 200.0]
    points = [(p, b0 + b1 * p + b2 * p**2) for p in ps]

    result = polynomial_fit.check_two_phase(points, degree=2)

    for (_, actual), fitted in zip(points, result.fitted):
        assert fitted == pytest.approx(actual, rel=1e-9)
    assert result.max_dev_pct == pytest.approx(0.0, abs=1e-6)
    assert result.qc.check_id == "cce_tp_fit_dev_pct"
    assert result.qc.severity == Severity.PASS
    assert len(result.coeffs) == 3


def test_fit_result_coeffs_and_scaling_reproduce_fitted_values():
    # Contract check (task brief): coeffs are in the SCALED basis, with
    # the scaling params carried alongside so callers can evaluate them
    # independently of `fitted`.
    a0, a1, a2, a3 = 0.9, 1.0e-5, -1.0e-9, 4.0e-14
    ps = [6000.0, 5000.0, 4000.0, 3000.0, 2000.0]
    points = [(p, a0 + a1 * p + a2 * p**2 + a3 * p**3) for p in ps]

    result = polynomial_fit.check_single_phase(points, degree=3)

    for (p, _), fitted in zip(points, result.fitted):
        p_scaled = (p - result.p_mean) / result.p_std
        reeval = sum(c * p_scaled**k for k, c in enumerate(result.coeffs))
        assert reeval == pytest.approx(fitted, rel=1e-9)


def test_check_single_phase_realistic_psia_range_does_not_blow_up():
    # Ill-conditioning guard: a degree-3 fit over the fixture's realistic
    # psia range (~500-7000) must not blow up (inf/nan/huge residual) --
    # this is exactly what centering/scaling P before the normal-equations
    # solve (see module docstring) is meant to prevent. Smooth data =
    # a mild cubic-ish trend plus a small bounded wobble a cubic can't
    # perfectly track, so a *small* nonzero residual is expected and fine;
    # a *blown-up* one (inf/nan or wildly large) would signal the
    # numerical-care guard failed.
    ps = [500.0 + 300.0 * i for i in range(23)]  # 500..7100 step 300
    points = [
        (p, 0.97 + 4.0e-5 * (7000.0 - p) - 1.2e-9 * (7000.0 - p) ** 2 + 6.0e-4 * math.sin(p / 173.0))
        for p in ps
    ]

    result = polynomial_fit.check_single_phase(points, degree=3)

    assert math.isfinite(result.max_dev_pct)
    assert math.isfinite(result.mean_dev_pct)
    assert all(math.isfinite(c) for c in result.coeffs)
    assert result.max_dev_pct < 5.0  # generous bound: "small", not blown up


# --- polynomial_fit: validation guards --------------------------------------


def test_check_single_phase_insufficient_points_raises():
    with pytest.raises(InputValidationError):
        polynomial_fit.check_single_phase([(100.0, 1.0), (200.0, 1.1)], degree=3)


def test_check_two_phase_insufficient_points_raises():
    with pytest.raises(InputValidationError):
        polynomial_fit.check_two_phase([(100.0, 1.0)], degree=2)


def test_check_single_phase_all_identical_pressure_raises():
    points = [(1000.0, v) for v in (1.0, 1.01, 0.99, 1.02)]
    with pytest.raises(InputValidationError):
        polynomial_fit.check_single_phase(points, degree=3)


def test_check_single_phase_too_few_distinct_pressures_raises():
    # 4 points (>= degree+1=4, passes the count guard) but only 2 DISTINCT
    # pressures -> the design matrix has rank <= 2 < 4, so X^T X is exactly
    # singular -- this exercises the solver's own pivot guard, distinct
    # from the p_std==0 (all-identical) guard above.
    points = [(100.0, 1.0), (100.0, 1.5), (100.0, 1.2), (200.0, 3.0)]
    with pytest.raises(InputValidationError):
        polynomial_fit.check_single_phase(points, degree=3)


def test_check_two_phase_zero_actual_value_raises():
    points = [(100.0, 0.0), (200.0, 1.0), (300.0, 2.0), (400.0, 3.0)]
    with pytest.raises(InputValidationError):
        polynomial_fit.check_two_phase(points, degree=2)


# --- polynomial_fit: grading bands (offline-computed, see module docstring) -

# OFFLINE DERIVATION (T6-fix pattern): thresholds cce_sp_fit_dev_pct =
# (0.05, 0.10)% are extremely tight, so hitting them with hand-picked round
# numbers isn't practical -- instead, a degree-1 override (analytically an
# ordinary 2-parameter line fit, same grading code path as the module's
# degree-3 default) was used with 3 points [(0,C),(1,C),(2,C+delta)] --
# baseline flat, one point perturbed by `delta` -- and `delta` was swept to
# find values landing cleanly inside each band, confirmed by calling
# polynomial_fit.check_single_phase directly during authoring (C=100.0):
#   delta = 0.15 -> max_dev_pct = 0.04999999999999716  (still PASS, <0.05)
#   delta = 0.18 -> max_dev_pct = 0.060000000000002274 (REVIEW, just over
#     the 0.05 edge -- too close to the boundary to use safely)
#   delta = 0.20 -> max_dev_pct = 0.06666666666666288  (REVIEW, comfortable
#     margin from both 0.05 and 0.10 edges -- used below)
#   delta = 0.30 -> max_dev_pct = 0.10000000000000853  (FAIL, but only ~1e-13
#     over the 0.10 edge -- too close to the boundary to use safely)
#   delta = 0.50 -> max_dev_pct = 0.1666666666666714   (FAIL, comfortable
#     margin above 0.10 -- used below)
def test_check_single_phase_review_band_offline_derived():
    points = [(0.0, 100.0), (1.0, 100.0), (2.0, 100.20)]
    result = polynomial_fit.check_single_phase(points, degree=1)
    assert result.max_dev_pct == pytest.approx(0.06666666666666288, rel=1e-9)
    assert result.qc.severity == Severity.REVIEW


def test_check_single_phase_fail_band_offline_derived():
    points = [(0.0, 100.0), (1.0, 100.0), (2.0, 100.50)]
    result = polynomial_fit.check_single_phase(points, degree=1)
    assert result.max_dev_pct == pytest.approx(0.1666666666666714, rel=1e-9)
    assert result.qc.severity == Severity.FAIL


# OFFLINE DERIVATION: cce_tp_fit_dev_pct = (1.0, 2.0)%, same construction
# (degree=1 override, 3 points, flat baseline C=10.0 + one perturbed
# point), swept and confirmed by calling polynomial_fit.check_two_phase
# directly during authoring:
#   delta = 0.30 -> max_dev_pct = 0.9999999999999963 (still PASS, <1.0)
#   delta = 0.45 -> max_dev_pct = 1.5000000000000036 (REVIEW, comfortable
#     margin from both 1.0 and 2.0 edges -- used below)
#   delta = 0.60 -> max_dev_pct = 2.0000000000000107 (FAIL, but only ~1e-14
#     over the 2.0 edge -- too close to the boundary to use safely)
#   delta = 0.70 -> max_dev_pct = 2.333333333333325  (FAIL, comfortable
#     margin above 2.0 -- used below)
def test_check_two_phase_review_band_offline_derived():
    points = [(0.0, 10.0), (1.0, 10.0), (2.0, 10.45)]
    result = polynomial_fit.check_two_phase(points, degree=1)
    assert result.max_dev_pct == pytest.approx(1.5000000000000036, rel=1e-9)
    assert result.qc.severity == Severity.REVIEW


def test_check_two_phase_fail_band_offline_derived():
    points = [(0.0, 10.0), (1.0, 10.0), (2.0, 10.70)]
    result = polynomial_fit.check_two_phase(points, degree=1)
    assert result.max_dev_pct == pytest.approx(2.333333333333325, rel=1e-9)
    assert result.qc.severity == Severity.FAIL


# --- monotonic_compressibility ----------------------------------------------


def test_monotonic_clean_zero_violations_pass():
    stages = [(1000.0, 5.0), (900.0, 6.0), (800.0, 6.0), (700.0, 8.0)]
    result = monotonic_compressibility.check(stages)
    assert result.violations == []
    assert result.qc.value == 0.0
    assert result.qc.severity == Severity.PASS
    assert result.qc.check_id == "cce_monotonic_violations"


def test_monotonic_one_violation_review():
    # 900->800 decreases (6.0 -> 5.5): one violation.
    stages = [(1000.0, 5.0), (900.0, 6.0), (800.0, 5.5), (700.0, 8.0)]
    result = monotonic_compressibility.check(stages)
    assert len(result.violations) == 1
    v = result.violations[0]
    assert (v.p_prev, v.c_prev, v.p_next, v.c_next) == (900.0, 6.0, 800.0, 5.5)
    assert result.qc.value == 1.0
    assert result.qc.severity == Severity.REVIEW


def test_monotonic_two_violations_fail():
    stages = [(1000.0, 6.0), (900.0, 5.0), (800.0, 7.0), (700.0, 6.5)]
    result = monotonic_compressibility.check(stages)
    assert len(result.violations) == 2
    assert result.qc.value == 2.0
    assert result.qc.severity == Severity.FAIL


def test_monotonic_skips_none_entries():
    stages = [
        (1000.0, 5.0),
        (950.0, None),
        (900.0, 6.0),
        (850.0, None),
        (800.0, 4.0),  # violation vs the last surviving entry (900, 6.0)
    ]
    result = monotonic_compressibility.check(stages)
    assert len(result.violations) == 1
    v = result.violations[0]
    assert (v.p_prev, v.c_prev, v.p_next, v.c_next) == (900.0, 6.0, 800.0, 4.0)


def test_monotonic_equal_consecutive_not_a_violation():
    stages = [(1000.0, 5.0), (900.0, 5.0), (800.0, 5.0)]
    result = monotonic_compressibility.check(stages)
    assert result.violations == []
    assert result.qc.severity == Severity.PASS


def test_monotonic_fewer_than_two_valid_points_zero_violations():
    assert monotonic_compressibility.check([]).qc.value == 0.0
    assert monotonic_compressibility.check([(1000.0, 5.0)]).qc.value == 0.0
    assert monotonic_compressibility.check([(1000.0, None)]).qc.value == 0.0


# --- rho_v_constancy ---------------------------------------------------------


def test_rho_v_exact_constant_zero_spread_pass():
    points = [(1.0, 10.0), (0.5, 20.0), (0.25, 40.0)]  # rho*V == 10.0 always
    result = rho_v_constancy.check(points)
    assert result.value == pytest.approx(0.0)
    assert result.severity == Severity.PASS
    assert result.check_id == "cce_rho_v_spread_pct"


def test_rho_v_review_band_hand_computed():
    # products = [100, 100, 100, 100.7]; mean = 400.7/4 = 100.175;
    # max|dev| = |100.7-100.175| = 0.525 -> spread = 0.525/100.175*100.
    points = [(1.0, 100.0), (1.0, 100.0), (1.0, 100.0), (1.0, 100.7)]
    products = [rho * v for rho, v in points]
    mean_product = sum(products) / len(products)
    expected = max(abs(p - mean_product) for p in products) / abs(mean_product) * 100.0

    result = rho_v_constancy.check(points)
    assert result.value == pytest.approx(expected, rel=1e-12)
    assert 0.5 < result.value <= 1.0
    assert result.severity == Severity.REVIEW


def test_rho_v_fail_band_hand_computed():
    # products = [100, 100, 100, 102]; mean = 402/4 = 100.5;
    # max|dev| = |102-100.5| = 1.5 -> spread = 1.5/100.5*100 ~ 1.4925%.
    points = [(1.0, 100.0), (1.0, 100.0), (1.0, 100.0), (1.0, 102.0)]
    products = [rho * v for rho, v in points]
    mean_product = sum(products) / len(products)
    expected = max(abs(p - mean_product) for p in products) / abs(mean_product) * 100.0

    result = rho_v_constancy.check(points)
    assert result.value == pytest.approx(expected, rel=1e-12)
    assert result.value > 1.0
    assert result.severity == Severity.FAIL


def test_rho_v_insufficient_points_raises():
    with pytest.raises(InputValidationError):
        rho_v_constancy.check([(1.0, 10.0)])


def test_rho_v_mean_zero_raises():
    with pytest.raises(InputValidationError):
        rho_v_constancy.check([(1.0, 1.0), (1.0, -1.0)])


# --- integration: real fixture engine output --------------------------------

WB = Path("tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx")


def _load_fixture_cce_results():
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["CCE Calculation"]

    stages = []
    row = 16
    while ws[f"A{row}"].value is not None:
        stages.append(
            CceStage(
                step=int(ws[f"A{row}"].value),
                p=float(ws[f"B{row}"].value),
                v_cell_cc=float(ws[f"C{row}"].value),
            )
        )
        row += 1

    inputs = CceInputs(
        t_res_f=float(ws["D6"].value),
        psat_visual=float(ws["D9"].value),
        bubble_point_step=int(ws["D10"].value),
        stages=tuple(stages),
        rho_at_psat_g_cc=float(ws["J10"].value),
        reservoir_p_psia=float(ws["D5"].value),
    )
    return calculate(inputs)


def test_cce_checks_over_real_fixture_are_well_formed():
    results = _load_fixture_cce_results()

    sp_points = [(s.p, s.rel_vol) for s in results.stages if s.p >= results.psat_from_data]
    tp_points = [(s.p, s.y_function) for s in results.stages if s.y_function is not None]
    mono_points = [(s.p, s.inst_compressibility_1e6_per_psi) for s in results.stages]
    rho_v_points = [
        (s.density_g_cc, s.rel_vol * results.v_sat_cc)
        for s in results.stages
        if s.density_g_cc is not None
    ]

    fit_sp = polynomial_fit.check_single_phase(sp_points)
    fit_tp = polynomial_fit.check_two_phase(tp_points)
    mono = monotonic_compressibility.check(mono_points)
    rho_v_qc = rho_v_constancy.check(rho_v_points)

    for qc in (fit_sp.qc, fit_tp.qc, mono.qc, rho_v_qc):
        assert isinstance(qc, QCResult)
        assert qc.severity in (Severity.PASS, Severity.REVIEW, Severity.FAIL)
        assert qc.value is not None and math.isfinite(qc.value)
        assert qc.message

    assert len(fit_sp.fitted) == len(sp_points)
    assert len(fit_tp.fitted) == len(tp_points)
    assert all(math.isfinite(c) for c in fit_sp.coeffs)
    assert all(math.isfinite(c) for c in fit_tp.coeffs)

    # Observed severities on this fixture run (informational only -- per
    # the task brief, NOT asserted: the sheet's own data may legitimately
    # land in REVIEW on these checks, especially the two resurrected/
    # proposal-status ones). Actual run, recorded here:
    #   cce_sp_fit_dev_pct:       PASS   (max_dev_pct = 0.0111%)
    #   cce_tp_fit_dev_pct:       PASS   (max_dev_pct = 0.4725%)
    #   cce_monotonic_violations: REVIEW (1 violation among the interior
    #     at/above-Psat compressibility points -- a real, single dip; the
    #     fixture's own instantaneous-compressibility series is not
    #     perfectly monotonic)
    #   cce_rho_v_spread_pct:     PASS   (spread_pct = 0.0% exactly -- the
    #     engine's density formula IS rho_i = rho_at_psat*v_sat/v_i, so
    #     rho_i*v_i is constant by construction; this check is tautological
    #     when fed the engine's own density output, and would only catch a
    #     genuine spread on independently-measured/imported density data)
