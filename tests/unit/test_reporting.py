"""Unit tests for pvt.reporting — report table builders and Excel export."""

import dataclasses

import openpyxl

from pvt.core.components import KATZ_FIROOZABADI as KF
from pvt.core.composition import CompositionStream
from pvt.core.sample import Sample
from pvt.experiments.flash.models import FlashResults
from pvt.experiments.flash.recombine import recombine_mass
from pvt.experiments.recombination.loading import LoadingInputs, plan_loading
from pvt.experiments.recombination.molar import GorBasis, molar_split
from pvt.qc.engine import QCResult, Severity
from pvt.reporting.excel_export import write_report
from pvt.reporting.tables import ReportRow, ReportTable, flash_tables, recombination_tables

SAMPLE = Sample(
    sample_id="SA-372",
    well="WELL-X",
    field_name="Upper Zakum",
    reservoir="Kharaib-2",
    depth_ft_md=9105.0,
    fluid_type="Black Oil",
    cylinder="RF1168636",
)


def _flash_results() -> FlashResults:
    return FlashResults(
        v_press_cc=20.8945,
        m_oil_g=13.71,
        v_gas_meas_cc=958.2037,
        v_gas_std_cc=940.5655,
        gas_density_std_g_cc=0.001404423,
        m_gas_g=1.32095,
        gor_cc_cc=59.6896,
        gor_scf_bbl=335.13,
        bo_flash=1.32600,
        shrinkage=0.754151,
        oil_density_60f_g_cc=0.870056,
        api=31.133,
    )


def _mass_recombination():
    oil = CompositionStream(library=KF, wt_pct={"C7": 80.0, "C1": 20.0})
    gas = CompositionStream(library=KF, wt_pct={"C1": 90.0, "C2": 10.0})
    return recombine_mass(13.71, 1.32095, oil, gas)


def _qc_results() -> list[QCResult]:
    return [
        QCResult(
            check_id="composition_sum",
            severity=Severity.PASS,
            value=0.1,
            threshold="review >0.5% / fail >2.0%",
            message="Sums to 100.1% (PASS)",
        ),
        QCResult(
            check_id="mass_balance_pct",
            severity=Severity.REVIEW,
            value=2.4,
            threshold="review >2.0% / fail >3.0%",
            message="Deviates 2.4% (REVIEW)",
        ),
        QCResult(
            check_id="gor_actual_vs_target_pct",
            severity=Severity.FAIL,
            value=12.0,
            threshold="review >5.0% / fail >10.0%",
            message="Deviates 12.0% (FAIL)",
        ),
    ]


# ---------------------------------------------------------------------------
# Step 1 (brief-verbatim): round-trip write -> load -> assert cell contents.
# ---------------------------------------------------------------------------


def test_round_trip_report(tmp_path):
    tables = [
        ReportTable(
            "Flash Results",
            [ReportRow("GOR", "335.13", "scf/bbl"), ReportRow("Bo", "1.3260", "vol/vol")],
        ),
        ReportTable("QC Summary", [ReportRow("composition_sum", "REVIEW", "")]),
    ]
    out = tmp_path / "report.xlsx"
    write_report(
        out,
        tables,
        title="Flash Separation Report",
        sample=Sample(
            sample_id="SA-372",
            well="WELL-X",
            field_name="Upper Zakum",
            reservoir="Kharaib-2",
            depth_ft_md=9105.0,
            fluid_type="Black Oil",
            cylinder="RF1168636",
        ),
    )
    ws = openpyxl.load_workbook(out).active
    text = [[c.value for c in row] for row in ws.iter_rows()]
    flat = [str(v) for row in text for v in row if v is not None]
    assert "Flash Separation Report" in flat and "SA-372" in flat
    assert "GOR" in flat and "335.13" in flat and "scf/bbl" in flat


# ---------------------------------------------------------------------------
# tables.flash_tables
# ---------------------------------------------------------------------------


def test_flash_tables_section_titles():
    tables = flash_tables(_flash_results(), _mass_recombination(), _qc_results())
    assert [t.title for t in tables] == ["Flash Results", "Whole Sample", "QC Summary"]


def test_flash_tables_flash_results_rows_cover_every_field():
    results = _flash_results()
    tables = flash_tables(results, _mass_recombination(), [])
    flash_table = next(t for t in tables if t.title == "Flash Results")
    assert len(flash_table.rows) == len(dataclasses.fields(results))
    values = [row.value for row in flash_table.rows]
    assert "335.13" in values  # gor_scf_bbl
    assert "1.3260" in values  # bo_flash


def test_flash_tables_whole_sample_rows():
    recomb = _mass_recombination()
    tables = flash_tables(_flash_results(), recomb, [])
    whole_sample = next(t for t in tables if t.title == "Whole Sample")
    labels = [row.label for row in whole_sample.rows]
    assert "Whole Sample MW" in labels
    mw_row = next(row for row in whole_sample.rows if row.label == "Whole Sample MW")
    assert mw_row.value == f"{recomb.mw_whole_sample:.2f}"


def test_flash_tables_every_qc_result_appears_as_a_row():
    qc = _qc_results()
    tables = flash_tables(_flash_results(), _mass_recombination(), qc)
    qc_table = next(t for t in tables if t.title == "QC Summary")
    assert len(qc_table.rows) == len(qc)
    for result, row in zip(qc, qc_table.rows):
        assert row.label == result.check_id
        assert row.value == result.severity.value
        assert row.unit == result.message


def test_flash_tables_qc_summary_empty_when_no_qc_results():
    tables = flash_tables(_flash_results(), _mass_recombination(), [])
    qc_table = next(t for t in tables if t.title == "QC Summary")
    assert qc_table.rows == []


# ---------------------------------------------------------------------------
# tables.recombination_tables
# ---------------------------------------------------------------------------


def _loading_inputs(target_oil_cc: float = 300.0) -> LoadingInputs:
    return LoadingInputs(
        cylinder_volume_cc=1000.0,
        target_oil_cc=target_oil_cc,
        oil_load_p_psig=100.0,
        oil_load_t_f=100.0,
        gas_load_p_psig=2000.0,
        gas_load_t_f=100.0,
        z_gas_load=0.85,
        sto_density_at_load_g_cc=0.83,
    )


def _molar_split():
    return molar_split(
        gor=850.0,
        basis=GorBasis.STOCK_TANK,
        shrinkage=1.0,
        sto_density_g_cc=0.83,
        sto_mw=210.0,
        gas_mw=22.0,
    )


def test_recombination_tables_section_titles():
    split = _molar_split()
    plan = plan_loading(_loading_inputs(), split, sto_density_60f=0.83, sto_mw=210.0)
    tables = recombination_tables(split, plan, _qc_results())
    assert [t.title for t in tables] == ["Molar Split", "Loading Plan", "QC Summary"]


def test_recombination_tables_molar_split_rows_cover_every_field():
    split = _molar_split()
    plan = plan_loading(_loading_inputs(), split, sto_density_60f=0.83, sto_mw=210.0)
    tables = recombination_tables(split, plan, [])
    split_table = next(t for t in tables if t.title == "Molar Split")
    assert len(split_table.rows) == len(dataclasses.fields(split))


def test_recombination_tables_loading_plan_fits_true():
    split = _molar_split()
    plan = plan_loading(
        _loading_inputs(target_oil_cc=300.0), split, sto_density_60f=0.83, sto_mw=210.0
    )
    assert plan.fits is True
    tables = recombination_tables(split, plan, [])
    loading_table = next(t for t in tables if t.title == "Loading Plan")
    fits_row = next(r for r in loading_table.rows if r.label == "Fits Cylinder")
    assert fits_row.value == "Yes"


def test_recombination_tables_loading_plan_fits_false():
    split = _molar_split()
    plan = plan_loading(
        _loading_inputs(target_oil_cc=990.0), split, sto_density_60f=0.83, sto_mw=210.0
    )
    assert plan.fits is False
    tables = recombination_tables(split, plan, [])
    loading_table = next(t for t in tables if t.title == "Loading Plan")
    fits_row = next(r for r in loading_table.rows if r.label == "Fits Cylinder")
    assert fits_row.value == "No"


def test_recombination_tables_every_qc_result_appears_as_a_row():
    split = _molar_split()
    plan = plan_loading(_loading_inputs(), split, sto_density_60f=0.83, sto_mw=210.0)
    qc = _qc_results()
    tables = recombination_tables(split, plan, qc)
    qc_table = next(t for t in tables if t.title == "QC Summary")
    assert len(qc_table.rows) == len(qc)
    for result, row in zip(qc, qc_table.rows):
        assert row.label == result.check_id
        assert row.value == result.severity.value
        assert row.unit == result.message


# ---------------------------------------------------------------------------
# excel_export.write_report — header style + severity fill spot checks
# ---------------------------------------------------------------------------


def test_write_report_header_style(tmp_path):
    tables = [ReportTable("Flash Results", [ReportRow("GOR", "335.13", "scf/bbl")])]
    out = tmp_path / "header.xlsx"
    write_report(out, tables, title="Flash Separation Report", sample=SAMPLE)

    ws = openpyxl.load_workbook(out).active
    title_cell = ws["A1"]
    assert title_cell.value == "Flash Separation Report"
    assert title_cell.font.bold is True
    assert title_cell.font.color.rgb == "00FFFFFF"
    assert title_cell.fill.fgColor.rgb == "0000205B"


def test_write_report_section_title_is_bold(tmp_path):
    tables = [ReportTable("Flash Results", [ReportRow("GOR", "335.13", "scf/bbl")])]
    out = tmp_path / "section.xlsx"
    write_report(out, tables, title="Flash Separation Report", sample=SAMPLE)

    ws = openpyxl.load_workbook(out).active
    section_cells = [
        cell
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "Flash Results"
    ]
    assert len(section_cells) == 1
    assert section_cells[0].font.bold is True


def test_write_report_severity_cells_filled(tmp_path):
    tables = [
        ReportTable(
            "QC Summary",
            [
                ReportRow("composition_sum", "PASS", "Sums to 100.1%"),
                ReportRow("mass_balance_pct", "REVIEW", "Deviates 2.4%"),
                ReportRow("gor_actual_vs_target_pct", "FAIL", "Deviates 12.0%"),
            ],
        )
    ]
    out = tmp_path / "severity.xlsx"
    write_report(out, tables, title="QC Report", sample=SAMPLE)

    ws = openpyxl.load_workbook(out).active
    fills = {
        cell.value: cell.fill.fgColor.rgb
        for row in ws.iter_rows()
        for cell in row
        if cell.value in ("PASS", "REVIEW", "FAIL")
    }
    assert fills == {
        "PASS": "0038A169",
        "REVIEW": "00DD9A0A",
        "FAIL": "00E53E3E",
    }


def test_write_report_handles_missing_depth(tmp_path):
    sample = dataclasses.replace(SAMPLE, depth_ft_md=None)
    tables = [ReportTable("Flash Results", [ReportRow("GOR", "335.13", "scf/bbl")])]
    out = tmp_path / "no_depth.xlsx"
    write_report(out, tables, title="Flash Separation Report", sample=sample)

    ws = openpyxl.load_workbook(out).active
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "N/A" in flat


def test_write_report_unit_column_wide_enough_for_qc_messages(tmp_path):
    # Demo-minor fix: the QC Summary section's unit column carries the full
    # check message (pvt.reporting.tables._qc_rows) -- must be wide enough
    # that a realistic message isn't visually clipped to a handful of
    # characters (the stored cell value is never truncated regardless, but
    # the column must still be usable without manual resizing).
    tables = [ReportTable("QC Summary", [ReportRow("hoffman_r2", "REVIEW", "message")])]
    out = tmp_path / "width.xlsx"
    write_report(out, tables, title="QC Report", sample=SAMPLE)
    ws = openpyxl.load_workbook(out).active
    assert ws.column_dimensions["C"].width >= 40


def test_write_report_accepts_str_path(tmp_path):
    tables = [ReportTable("Flash Results", [ReportRow("GOR", "335.13", "scf/bbl")])]
    out = str(tmp_path / "str_path.xlsx")
    write_report(out, tables, title="Flash Separation Report", sample=SAMPLE)
    assert openpyxl.load_workbook(out).active["A1"].value == "Flash Separation Report"
