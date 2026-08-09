"""Golden/integration tests for the Flash v6.1 Excel importer
(`pvt.io.excel_import.flash_v61`).

`test_import_then_calculate_reproduces_workbook` and `test_wrong_file_rejected`
are transcribed verbatim from the task-7 brief. The remaining tests are
supplementary: they cross-check the full cell map (not just the three
calc.py outputs the brief's golden test exercises) against two already-golden
fixtures — `tests/unit/experiments/test_flash_validate.SA372` (Task 1) and
`tests/fixtures/sa372_flash.py` (Task 3) — and cover the negative-composition
guard and the shifted-header wrong-file guard, both needed for the
--cov-fail-under=100 branch-coverage gate.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from pvt.core.components import KATZ_FIROOZABADI as KF
from pvt.core.exceptions import InputValidationError
from pvt.experiments.flash.calc import calculate
from pvt.io.excel_import.flash_v61 import read
from tests.fixtures import sa372_flash as fx
from tests.unit.experiments.test_flash_validate import SA372

WB = Path("tests/fixtures/workbooks/ADRIC_Flash_Separation_Calc_v6.1.xlsx")
LIVEOIL_WB = Path("tests/fixtures/workbooks/ADRIC_LiveOil_Preparation_Calc_v4.1.xlsx")


def test_import_then_calculate_reproduces_workbook():
    imp = read(WB)
    r = calculate(imp.volumetrics)
    assert r.gor_scf_bbl == pytest.approx(335.13, abs=0.01)
    assert r.bo_flash == pytest.approx(1.32600, abs=1e-5)
    assert r.api == pytest.approx(31.133, abs=0.001)
    assert imp.sample.sample_id == "SA-372"
    assert imp.oil_stream.raw_mol_sum() == pytest.approx(100.0, abs=0.5)


def test_wrong_file_rejected():
    with pytest.raises(InputValidationError):
        read(LIVEOIL_WB)


def test_volumetrics_cell_map_matches_sa372_fixture():
    # Cross-check against Task 1's golden SA372 FlashVolumetrics (dataclass
    # equality checks all 13 fields at once, not just the 3 calc.py exposes).
    imp = read(WB)
    assert imp.volumetrics == SA372


def test_composition_cell_map_matches_sa372_flash_fixture():
    # Cross-check against Task 3's generated fixture (all 52 rows x 4 columns,
    # including the 3-code alias map), independently of this importer.
    imp = read(WB)
    assert imp.gas_stream.mol_pct == fx.GAS_MOL_PCT
    assert imp.gas_stream.wt_pct == fx.GAS_WT_PCT
    assert imp.oil_stream.mol_pct == fx.OIL_MOL_PCT
    assert imp.oil_stream.wt_pct == fx.OIL_WT_PCT


def test_sample_metadata_mapped():
    imp = read(WB)
    s = imp.sample
    assert s.well == "WELL-X"
    assert s.field_name == "Upper Zakum"
    assert s.client == "ADNOC Onshore"
    assert s.cylinder == "RF1168636"
    assert s.project == "ADRIC-PVT-2026-014"
    assert s.fluid_type == "Black Oil"
    assert s.depth_ft_md == pytest.approx(9105.0)


def _doctored_workbook(
    tmp_path: Path,
    *,
    header_cell: tuple[str, str] | None = None,
    composition_override: tuple[int, str, float | None] | None = None,
    blank_cell: str | None = None,
) -> Path:
    """Build a minimal but structurally valid Volumetrics_Master workbook,
    in-memory, for negative-path tests that must not touch the real fixture
    files. Mirrors the real template's layout (title, row-40 header, the
    52-code composition block in KF order) closely enough to reach the
    code path under test.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Volumetrics_Master"
    ws["A1"] = "ATMOSPHERIC FLASH SEPARATION - WATER PUMP METHOD (v6.1)"

    metadata = {
        "B5": "ADNOC Onshore", "E5": 256, "H5": "SA-372",
        "B6": "WELL-X / Upper Zakum", "E6": 3939, "H6": "RF1168636",
        "B7": 9105, "E7": 1156, "H7": "M. Athif",
        "B8": "2026-04-12", "E8": "Black Oil", "H8": "ADRIC-PVT-2026-014",
    }
    for addr, value in metadata.items():
        ws[addr] = value

    volumetrics = {
        "B11": 1, "E11": 1, "B12": 50.0, "E12": 70.8945, "B14": 15.7576,
        "B15": 100.0, "E15": 113.71, "B17": 1, "B18": 500.0, "E18": 1458.2037,
        "B20": 20.0, "E17": 1013.25, "E20": 1, "B21": 1012.25, "E21": 1.146,
    }
    for addr, value in volumetrics.items():
        ws[addr] = value

    header = {
        "A40": "#", "B40": "Code", "C40": "Component", "D40": "MW (KF)",
        "E40": "Gas Mol% (INPUT)", "F40": "Gas Wt% (INPUT)",
        "G40": "Oil Mol% (INPUT)", "H40": "Oil Wt% (INPUT)",
    }
    for addr, value in header.items():
        ws[addr] = value

    for i, code in enumerate(KF.codes):
        row = 41 + i
        ws[f"B{row}"] = code
        # Row 0 seeded non-zero (not the row any current test blanks/negates)
        # so a doctored workbook always yields a non-empty CompositionStream,
        # even when a test blanks/negates one specific other row's cell.
        seed = 50.0 if i == 0 else 0
        ws[f"E{row}"] = seed
        ws[f"F{row}"] = seed
        ws[f"G{row}"] = seed
        ws[f"H{row}"] = seed

    if header_cell is not None:
        addr, value = header_cell
        ws[addr] = value
    if composition_override is not None:
        row_offset, col, value = composition_override
        ws[f"{col}{41 + row_offset}"] = value
    if blank_cell is not None:
        ws[blank_cell] = None

    path = tmp_path / "doctored.xlsx"
    wb.save(str(path))
    return path


def test_negative_composition_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, composition_override=(5, "G", -1.0))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "negative" in str(exc_info.value).lower()


def test_unexpected_header_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, header_cell=("B40", "WRONG"))
    with pytest.raises(InputValidationError):
        read(path)


def test_blank_composition_cell_is_treated_as_absent(tmp_path):
    """A blank (None) composition cell must not crash the `value < 0` sign
    check with a raw TypeError -- it is treated as absent, same as an
    explicit zero, and the component is simply missing from that stream's
    dict rather than blocking the import."""
    path = _doctored_workbook(tmp_path, composition_override=(5, "G", None))
    imp = read(path)
    missing_code = KF.codes[5]
    assert missing_code not in imp.oil_stream.mol_pct


def test_blank_required_cell_raises_typed_error_naming_the_cell(tmp_path):
    """A blank REQUIRED scalar cell (B14, v_sto_cc) must raise a typed
    InputValidationError naming the exact cell address rather than crashing
    `float(None)` with a raw TypeError."""
    path = _doctored_workbook(tmp_path, blank_cell="B14")
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "Volumetrics_Master!B14" in str(exc_info.value)
    assert "blank or non-numeric" in str(exc_info.value)


def test_blank_sample_depth_cell_raises_typed_error(tmp_path):
    """Same guard, covering _read_sample's own `_num` call (B7, sampling
    depth) rather than _read_volumetrics's."""
    path = _doctored_workbook(tmp_path, blank_cell="B7")
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "Volumetrics_Master!B7" in str(exc_info.value)
