"""
tests/golden/test_cce_workbook.py — CCE engine vs the dissected workbook's
cached cells (GOLDEN-INTEGRITY: expected values are read live from the
fixture with openpyxl, never hand-typed from a digest).

Fixture: tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx
Sheets used: "CCE Calculation" (stage table + auto-calculated params J6:J11)
and "Mean Compressibility" (D4, H7/H8, D9).

Cell map (confirmed by loading the fixture with both data_only=True and
data_only=False -- see also pvt/experiments/cce/models.py and
tests/unit/experiments/test_cce_validate.py):
    D6 = Temperature (F); D9 = Visual Bubble Point (psig); D10 = Bubble
    Point Step # (1-based). Stage table rows 16-55: A=step, B=P
    (psig-as-entered), C=entered cell volume (cc). D/E/F/G (Relative
    Volume / Density / Inst. Compressibility / Y-Function) are all
    DERIVED (formula cells) -- this test recomputes them with the engine
    and compares against the sheet's own cached results, it never reads
    D/E/F/G as engine input.

J-column auto-calculated parameters (rows 6-11):
    J6 = C16 (volume at the first/"Working Pressure" stage)
    J7 = D8*J6 = "Sample Mass (g)" -- D8 ("Measured HPHT Density") times
         the volume at the first stage; mass held constant while
         single-phase (mass conservation).
    J8 = V_sat (INDEX into C at the bubble row)
    J9 = Psat from Data (INDEX into B at the bubble row)
    J10 = J7/J8 = "Density at Psat" (g/cm3) -- this is the quantity that
          corresponds to CceInputs.rho_at_psat_g_cc (NOT D8, which is
          measured at the first-stage/"Working Pressure" condition, a
          different, generally-higher pressure than Psat). Column E's
          formula IF(A<=D10, J7/C, "") is algebraically identical to
          rho_at_psat * v_sat / v_i once rho_at_psat := J10 = J7/J8
          (J7/C_i = (J7/J8)*J8/C_i = rho_at_psat*v_sat/v_i), so wiring
          CceInputs.rho_at_psat_g_cc = J10's cached value reproduces
          column E exactly -- verified below at rel=1e-9.
    J11/K11 = |D9-J9| consistency gate (psat_consistency_ok).

F-column (instantaneous compressibility) stencil, read with
data_only=False: general rows (17-54) use a genuine central difference,
F_i = (C_{i+1}-C_{i-1})/C_i/(B_{i-1}-B_{i+1})*1e6, gated on A_i<=D10 (at
or above Psat). This central form needs both neighbours, so it cannot
apply to the very first stage (row 16: no row 15 to reference) -- the
sheet instead special-cases row 16 with a one-sided forward-difference
formula, (C17-C16)/C16/(B16-B17)*1e6. The engine implements ONLY the
central-difference method (per the Task 2 brief); row 16 is therefore
structurally a boundary the engine cannot evaluate (no i-1) and returns
None there, diverging from the sheet's one-sided F16 -- this is not
compared. At the OTHER end of the at/above-Psat gate, row 35 (the bubble
row, A35=20=D10) is where the sheet's central-difference stencil reaches
into row 36 (two-phase, A36=21>D10): F35 is contaminated (cached 85.17;
ledger D-019, docs/excel-deviations.md). The engine excludes both the
bubble row and this boundary row structurally (central diff needs
i-1>=0 and i+1 at/above Psat), so only rows 17-34 (interior, steps 2-19)
are compared against the sheet 1:1.

Mean Compressibility sheet:
    D4 = Range 1 (row16 -> row19), correct two-point form -- direct,
         uncontested golden check of the engine's two-point helper.
    H7 = 'CCE Calculation'!D5 ("Reservoir Pressure", a SEPARATE lab input
         from D7 "Working Pressure" = row16's pressure). H8 = "Mean Comp
         from Reservoir Pressure to Psat", INDEX/MATCHing H7 to the CCE
         table (which lands on row 23, NOT row 16 -- D5=3938.73 equals
         B23, not B16=D7=7014.73) then applying the two-point form with
         the numerator operands FLIPPED (ledger D-020): cached -12.365
         (negative). D9 duplicates the identical row23->row35 range with
         the CORRECT operand order and is the positive, correct-physics
         counterpart: cached +12.365433539344826 = abs(H8 cached),
         verified below.

ROUND 2 RESOLUTION (controller adjudication of the round-1 finding this
module originally reported): CceInputs gained an optional
`reservoir_p_psia` field (sheet D5). `_load()` below wires it from the
fixture, and `calc.py`'s `res_to_psat` replicates `H8`'s own
`MATCH(D5, B:B, -1)` anchor-row selection over the stage table (smallest
P still >= D5, landing on row 23 here) before applying the two-point
formula with the correct operand order -- so `res_to_psat` now
reproduces `abs(H8 cached)`/`D9 cached` exactly (verified below). The
always-available `first_stage_to_psat` key (row16->row35, the round-1
"res_to_psat") is unaffected and kept for backward compatibility; it is
still, correctly, a different number (10.397403562555192) since it
anchors to a different row (16, not 23) -- that's a self-consistency
arithmetic check against the same shared helper, not a workbook cell,
and is labeled as such below.
"""

from pathlib import Path

import openpyxl
import pytest

from pvt.experiments.cce.calc import calculate, mean_compressibility_1e6_per_psi
from pvt.experiments.cce.models import CceInputs, CceStage

WB = Path("tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx")


def _load() -> tuple[CceInputs, dict]:
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["CCE Calculation"]
    mc = wb["Mean Compressibility"]

    stages = []
    row = 16
    while ws[f"A{row}"].value is not None:
        stages.append(
            CceStage(
                step=int(ws[f"A{row}"].value),
                p=float(ws[f"B{row}"].value),
                v_cell_cc=float(ws[f"C{row}"].value),
            )
        )
        row += 1

    rho_at_psat = float(ws["J10"].value)  # "Density at Psat" -- see module docstring
    reservoir_p = float(ws["D5"].value)  # "Reservoir Pressure" -- anchors res_to_psat
    inputs = CceInputs(
        t_res_f=float(ws["D6"].value),
        psat_visual=float(ws["D9"].value),
        bubble_point_step=int(ws["D10"].value),
        stages=tuple(stages),
        rho_at_psat_g_cc=rho_at_psat,
        reservoir_p_psia=reservoir_p,
    )

    cached = {
        "D": [ws[f"D{r}"].value for r in range(16, 56)],
        "E": [ws[f"E{r}"].value for r in range(16, 56)],
        "F": [ws[f"F{r}"].value for r in range(16, 56)],
        "G": [ws[f"G{r}"].value for r in range(16, 56)],
        "J8": float(ws["J8"].value),
        "J9": float(ws["J9"].value),
        "J11": float(ws["J11"].value),
        "MC_D4": float(mc["D4"].value),  # Mean Compressibility!D4 (Range 1)
        "MC_H8": float(mc["H8"].value),  # Mean Compressibility!H8 (ledger D-020, flipped)
        "MC_D9": float(mc["D9"].value),  # Mean Compressibility!D9 (correct-order duplicate)
    }
    return inputs, cached


INPUTS, CACHED = _load()
RESULTS = calculate(INPUTS)
BUBBLE_IDX = INPUTS.bubble_point_step - 1  # 0-based; 19 for this fixture


def test_fixture_shape_sanity():
    assert len(INPUTS.stages) == 40
    assert INPUTS.bubble_point_step == 20
    assert INPUTS.psat_visual == pytest.approx(1155.73)


def test_v_sat_and_psat_from_data_match_j8_j9():
    assert RESULTS.v_sat_cc == pytest.approx(CACHED["J8"], rel=1e-9)
    assert RESULTS.psat_from_data == pytest.approx(CACHED["J9"], rel=1e-9)
    assert RESULTS.psat_from_data == INPUTS.stages[BUBBLE_IDX].p


def test_psat_consistency_matches_j11_k11_gate():
    # J11 = |D9-J9| = 0 < 10 -> K11 = "OK"
    assert CACHED["J11"] == pytest.approx(0.0, abs=1e-9)
    assert RESULTS.psat_consistency_ok is True


def test_relative_volume_all_40_rows():
    for idx, stage_result in enumerate(RESULTS.stages):
        assert stage_result.rel_vol == pytest.approx(CACHED["D"][idx], rel=1e-9)


def test_density_above_and_at_psat_matches_column_e():
    for idx in range(0, BUBBLE_IDX + 1):
        assert RESULTS.stages[idx].density_g_cc == pytest.approx(
            CACHED["E"][idx], rel=1e-6
        )


def test_density_none_below_psat():
    for idx in range(BUBBLE_IDX + 1, len(RESULTS.stages)):
        assert RESULTS.stages[idx].density_g_cc is None


def test_instantaneous_compressibility_interior_rows_match_column_f():
    # Interior rows only: idx 1..(BUBBLE_IDX-1) i.e. steps 2-19 (rows 17-34).
    # idx0 (row16, boundary, no i-1) and idx==BUBBLE_IDX (row35, bubble,
    # straddles into two-phase) are excluded -- see module docstring.
    for idx in range(1, BUBBLE_IDX):
        assert RESULTS.stages[idx].inst_compressibility_1e6_per_psi == pytest.approx(
            CACHED["F"][idx], rel=1e-6
        )


def test_instantaneous_compressibility_boundary_and_bubble_row_are_none():
    assert RESULTS.stages[0].inst_compressibility_1e6_per_psi is None
    assert RESULTS.stages[BUBBLE_IDX].inst_compressibility_1e6_per_psi is None
    # ledger D-019: the sheet's own F35 is contaminated by the two-phase
    # neighbour, cached ~85.17 -- confirms *why* it must be excluded, not
    # just *that* it differs.
    assert CACHED["F"][BUBBLE_IDX] == pytest.approx(85.16667974300425, rel=1e-6)


def test_instantaneous_compressibility_none_below_psat():
    for idx in range(BUBBLE_IDX + 1, len(RESULTS.stages)):
        assert RESULTS.stages[idx].inst_compressibility_1e6_per_psi is None


def test_y_function_below_psat_matches_column_g():
    for idx in range(BUBBLE_IDX + 1, len(RESULTS.stages)):
        assert RESULTS.stages[idx].y_function == pytest.approx(
            CACHED["G"][idx], rel=1e-6
        )


def test_y_function_none_at_and_above_psat():
    for idx in range(0, BUBBLE_IDX + 1):
        assert RESULTS.stages[idx].y_function is None


def test_mean_compressibility_range1_matches_sheet_d4():
    # Mean Compressibility!D4 = Range 1: row16 (step1) -> row19 (step4).
    # CACHED["MC_D4"] is read live in _load() (GOLDEN-INTEGRITY: never
    # hand-typed from a digest).
    row16, row19 = INPUTS.stages[0], INPUTS.stages[3]
    c = mean_compressibility_1e6_per_psi(
        v_i=row16.v_cell_cc, p_i=row16.p, v_f=row19.v_cell_cc, p_f=row19.p
    )
    assert c == pytest.approx(CACHED["MC_D4"], rel=1e-6)


def test_res_to_psat_matches_abs_h8_and_d9():
    # ledger D-020: Mean Compressibility!H8 flips the numerator operands
    # for the "Reservoir Pressure -> Psat" range and returns a negative
    # value; D9 duplicates the identical row23->row35 range with the
    # correct operand order and is the positive counterpart, = abs(H8).
    # CACHED["MC_H8"]/CACHED["MC_D9"] are read live in _load()
    # (GOLDEN-INTEGRITY: never hand-typed from a digest). Row 23 is
    # located generically here (independently of calc.py) to prove the
    # *formula*: it's the stage whose pressure matches D5 ("Reservoir
    # Pressure"), already read live into INPUTS.reservoir_p_psia.
    reservoir_p = INPUTS.reservoir_p_psia
    assert reservoir_p is not None
    reservoir_stage = next(s for s in INPUTS.stages if s.p == reservoir_p)
    bubble_stage = INPUTS.stages[BUBBLE_IDX]

    c = mean_compressibility_1e6_per_psi(
        v_i=reservoir_stage.v_cell_cc,
        p_i=reservoir_stage.p,
        v_f=bubble_stage.v_cell_cc,
        p_f=bubble_stage.p,
    )
    assert c == pytest.approx(CACHED["MC_D9"], rel=1e-6)
    assert c == pytest.approx(abs(CACHED["MC_H8"]), rel=1e-6)

    # ROUND 2: the engine's own res_to_psat (reservoir_p_psia=D5, wired
    # via _load()) now reproduces this exactly through calc.py's own
    # MATCH(-1)-equivalent anchor selection -- not just the raw helper
    # applied to a manually-located row.
    assert RESULTS.mean_compressibility_1_psi["res_to_psat"] == pytest.approx(
        CACHED["MC_D9"], rel=1e-6
    )
    assert RESULTS.mean_compressibility_1_psi["res_to_psat"] == pytest.approx(
        abs(CACHED["MC_H8"]), rel=1e-6
    )


def test_first_stage_to_psat_is_a_self_consistent_arithmetic_check():
    # first_stage_to_psat (row16/"Working Pressure" -> row35/bubble) has
    # no matching cached cell in this workbook (H8/D9 anchor to D5's
    # "Reservoir Pressure", row23, not row16 -- see res_to_psat above).
    # This is therefore a self-consistency check: the engine's own value
    # against an independent call to the same shared helper on the same
    # row pair -- NOT a comparison against any workbook cell.
    row16, bubble = INPUTS.stages[0], INPUTS.stages[BUBBLE_IDX]
    expected = mean_compressibility_1e6_per_psi(
        v_i=row16.v_cell_cc, p_i=row16.p, v_f=bubble.v_cell_cc, p_f=bubble.p
    )
    assert RESULTS.mean_compressibility_1_psi["first_stage_to_psat"] == pytest.approx(
        expected, rel=1e-9
    )
    # Sanity: distinct from the reservoir-row-anchored res_to_psat/D9/H8
    # value (read live, not hand-typed -- see CACHED["MC_D9"] above) --
    # documents that these are two legitimately different quantities,
    # not a bug in either.
    assert RESULTS.mean_compressibility_1_psi["first_stage_to_psat"] != pytest.approx(
        CACHED["MC_D9"], rel=1e-3
    )
