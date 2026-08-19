"""
tests/golden/test_psat_breakpoint.py — Phase 3a Task 4: Psat-breakpoint QC
(`pvt/qc/checks/psat_breakpoint.py`), a NEW capability dissected in-task from
`tests/fixtures/workbooks/Bubble_Dew_Point_QC_Tool_Final.xlsx`.

GOLDEN-INTEGRITY (same convention as `tests/golden/test_cce_workbook.py`):
the golden test below reads every expected number LIVE from the fixture with
openpyxl (both the raw data-entry cells and the cached formula results,
`data_only=True`) -- nothing is hand-typed from a digest. It even derives the
`split_at` pressure threshold generically from the sheet's own `K6` "Above
points used" cell rather than hand-typing a pressure value, so the whole
comparison is generator-driven.

The remaining tests use hand-constructed synthetic data (closed-form exact
lines, so the algebra is trivial to verify independently -- no offline
numerical sweep needed, unlike Task 3's polynomial-fit band tests) and a
final integration test against the real CCE fixture engine output (well-
formedness only, no severity asserted, per the task brief).
"""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pytest

from pvt.core.exceptions import InputValidationError
from pvt.experiments.cce.calc import calculate
from pvt.experiments.cce.models import CceInputs, CceStage
from pvt.qc.checks import psat_breakpoint
from pvt.qc.engine import QCResult, Severity, ThresholdRegistry

# --- registry / exports plumbing --------------------------------------------


def test_registry_has_new_psat_breakpoint_key():
    reg = ThresholdRegistry()
    assert reg.get("psat_breakpoint_vs_visual_psi") == (10.0, 25.0)


def test_checks_package_exports_psat_breakpoint():
    import pvt.qc.checks as checks

    assert "psat_breakpoint" in checks.__all__


# --- golden: dissected fixture vs live-read cached cells --------------------

WB = Path("tests/fixtures/workbooks/Bubble_Dew_Point_QC_Tool_Final.xlsx")


def test_check_reproduces_fixture_cached_intersection():
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["BP_DP_QC_Final"]

    # Read every data-entry row (14-25: #, Use, Pressure psig, Temp, QC Value)
    # live, in sheet row order -- never hand-typed.
    points = []
    row = 14
    while ws[f"A{row}"].value is not None:
        if int(ws[f"B{row}"].value) == 1:  # Use flag
            points.append((float(ws[f"C{row}"].value), float(ws[f"E{row}"].value)))
        row += 1

    # K6 = "Above points used" -- the sheet's own row-count region boundary
    # (see F-column formula `IF(A<=$K$6,"Above Point",...)`). The sheet
    # splits by ROW ORDER/COUNT, not by comparing pressure against a fixed
    # threshold cell (there isn't one) -- confirmed by dissection, see the
    # module docstring's "split semantics" note and workbook-defect-review.md
    # row BD1. Since the fixture's rows are pre-sorted descending by
    # pressure, a pressure-threshold split (this module's actual contract)
    # midway between the last above-row and first below-row reproduces the
    # SAME partition the sheet's row-count split produces -- derived here
    # generically from K6, not hand-typed.
    above_count = int(ws["K6"].value)
    split_at = (points[above_count - 1][0] + points[above_count][0]) / 2.0

    result = psat_breakpoint.check(points, split_at)

    assert result.slope_above == pytest.approx(float(ws["P4"].value), rel=1e-9)
    assert result.intercept_above == pytest.approx(float(ws["P5"].value), rel=1e-9)
    assert result.slope_below == pytest.approx(float(ws["P6"].value), rel=1e-9)
    assert result.intercept_below == pytest.approx(float(ws["P7"].value), rel=1e-9)
    assert result.psat_estimate == pytest.approx(float(ws["P8"].value), rel=1e-9)
    assert result.parallel_warning is False
    # Cross-check against the sheet's own narrative verdict cells (K10=
    # "OK", P10="Intersection calculated" -- both agree the fit is sane,
    # consistent with parallel_warning being False here).
    assert ws["K10"].value == "OK"
    assert ws["P10"].value == "Intersection calculated"
    assert result.qc.severity == Severity.PASS  # visual_psat=None -> trivial PASS
    assert result.qc.message == "no visual Psat provided; estimate only"


# --- synthetic exact-intersection dataset (two known lines, rel 1e-9) -------


def test_check_reconstructs_exact_intersection_of_two_known_lines():
    # Above line: y = 100 - 2p (m=-2, b=100). Below line: y = -50 + 3p
    # (m=3, b=-50). Intersection: 100-2p = -50+3p -> 150 = 5p -> p* = 30,
    # y* = 100-2*30 = 40. Points are placed away from p*=30 on each side of
    # split_at=20 so the split is unambiguous; the algebra above is exact
    # and independent of which points are chosen, since they lie exactly on
    # each line.
    above = [(25.0, 100 - 2 * 25.0), (30.0, 100 - 2 * 30.0), (35.0, 100 - 2 * 35.0), (40.0, 100 - 2 * 40.0)]
    below = [(5.0, -50 + 3 * 5.0), (10.0, -50 + 3 * 10.0), (15.0, -50 + 3 * 15.0), (18.0, -50 + 3 * 18.0)]

    result = psat_breakpoint.check(above + below, split_at=20.0)

    assert result.slope_above == pytest.approx(-2.0, rel=1e-9)
    assert result.intercept_above == pytest.approx(100.0, rel=1e-9)
    assert result.slope_below == pytest.approx(3.0, rel=1e-9)
    assert result.intercept_below == pytest.approx(-50.0, rel=1e-9)
    assert result.psat_estimate == pytest.approx(30.0, rel=1e-9)
    assert result.parallel_warning is False


# --- split semantics: p >= split_at is "above", p < split_at is "below" ----


def test_check_split_is_inclusive_above_exclusive_below():
    # A point exactly AT split_at must land in the above segment, not below
    # -- per the brief's confirmed convention (p >= split_at above).
    above = [(20.0, 1.0), (25.0, 1.0), (30.0, 1.0)]  # slope 0, intercept 1
    below = [(5.0, 2.0), (10.0, 2.0)]  # slope 0, intercept 2 (parallel, fine -- not under test)
    result = psat_breakpoint.check(above + below, split_at=20.0)
    # If the point (20.0, 1.0) had instead been classified "below", the
    # above segment would only have 2 points (25, 30) and the fit would be
    # identical (still slope 0) -- so to make the boundary observable, use
    # an above segment where dropping the boundary point CHANGES the fit.
    above2 = [(20.0, 5.0), (25.0, 1.0), (30.0, 1.0)]
    below2 = [(5.0, 2.0), (10.0, 2.0), (15.0, 2.0)]
    result2 = psat_breakpoint.check(above2 + below2, split_at=20.0)
    # With (20.0, 5.0) included in "above", the above-fit is pulled off
    # slope 0 -- confirms the boundary point participated in the above fit.
    assert result2.slope_above != pytest.approx(0.0, abs=1e-9)
    assert result.slope_above == pytest.approx(0.0, abs=1e-9)


# --- parallel-trend guard ----------------------------------------------------


def test_check_near_parallel_forces_review_regardless_of_close_visual_psat():
    # Above: y = p exactly (m=1, b=0) through (10,10),(20,20),(30,30).
    # Below: y = 1.0005*p exactly (m=1.0005, b=0) through (1,1.0005),
    # (2,2.001),(3,3.0015). |m1-m2| = 0.0005; scale = max(1,1.0005) =
    # 1.0005; 1e-3*scale = 0.0010005 -> 0.0005 < 0.0010005 -> near-parallel.
    # Intersection: p*(1) = p*(1.0005) + 0 - 0 -> only exactly true at p=0,
    # so b_above=b_below=0 -> psat_estimate = (0-0)/(1-1.0005) = 0.0.
    above = [(10.0, 10.0), (20.0, 20.0), (30.0, 30.0)]
    below = [(1.0, 1.0005), (2.0, 2.001), (3.0, 3.0015)]
    # visual_psat=0.0 is EXACTLY the estimate -- would trivially PASS a
    # normal distance grade -- but the parallel guard must override that.
    result = psat_breakpoint.check(above + below, split_at=5.0, visual_psat=0.0)

    assert result.parallel_warning is True
    assert result.psat_estimate == pytest.approx(0.0, abs=1e-6)
    assert result.qc.severity == Severity.REVIEW
    assert "parallel" in result.qc.message.lower()
    assert "ill-conditioned" in result.qc.message.lower()


def test_check_exactly_parallel_slopes_returns_nan_estimate_and_review():
    # Both segments have IDENTICAL slope and intercept -> the lines never
    # cross (or are literally the same line) -> the intersection formula's
    # denominator (m_above - m_below) is exactly zero -> psat_estimate is
    # defined as nan (never raises -- matches the sheet's own IFERROR(...,"")
    # blanking behavior, expressed as a numeric sentinel instead of a
    # string so downstream numeric code doesn't need to special-case it).
    above = [(10.0, 1.0), (20.0, 1.0), (30.0, 1.0)]  # slope 0, intercept 1
    below = [(1.0, 1.0), (2.0, 1.0), (3.0, 1.0)]  # slope 0, intercept 1 -- identical line
    result = psat_breakpoint.check(above + below, split_at=5.0)

    assert result.parallel_warning is True
    assert math.isnan(result.psat_estimate)
    assert result.qc.severity == Severity.REVIEW


# --- banded distance cases (offline-computed: exact closed-form here) ------

# Above line: y = 1000 - p (m=-1, b=1000). Below line: y = p - 1000 (m=1,
# b=-1000). Intersection: 1000-p = p-1000 -> 2000 = 2p -> p* = 1000 exactly
# (closed form, no numerical sweep needed -- both lines are exactly linear
# and the points below lie exactly on them).
_BAND_ABOVE = [(1100.0, 1000 - 1100.0), (1150.0, 1000 - 1150.0), (1200.0, 1000 - 1200.0), (1250.0, 1000 - 1250.0)]
_BAND_BELOW = [(700.0, 700.0 - 1000), (750.0, 750.0 - 1000), (800.0, 800.0 - 1000), (850.0, 850.0 - 1000)]


def _band_result(visual_psat: float) -> psat_breakpoint.BreakpointResult:
    return psat_breakpoint.check(_BAND_ABOVE + _BAND_BELOW, split_at=1000.0, visual_psat=visual_psat)


def test_check_visual_psat_within_review_band_is_pass():
    # |1000 - 1005| = 5 psi <= 10.0 -> PASS.
    result = _band_result(visual_psat=1005.0)
    assert result.psat_estimate == pytest.approx(1000.0, rel=1e-9)
    assert result.qc.value == pytest.approx(5.0, rel=1e-9)
    assert result.qc.severity == Severity.PASS


def test_check_visual_psat_within_fail_band_is_review():
    # |1000 - 1015| = 15 psi -- between 10.0 and 25.0 -> REVIEW.
    result = _band_result(visual_psat=1015.0)
    assert result.qc.value == pytest.approx(15.0, rel=1e-9)
    assert result.qc.severity == Severity.REVIEW


def test_check_visual_psat_beyond_fail_band_is_fail():
    # |1000 - 1030| = 30 psi > 25.0 -> FAIL.
    result = _band_result(visual_psat=1030.0)
    assert result.qc.value == pytest.approx(30.0, rel=1e-9)
    assert result.qc.severity == Severity.FAIL


# --- None-visual path --------------------------------------------------------


def test_check_none_visual_psat_is_trivial_pass():
    result = psat_breakpoint.check(_BAND_ABOVE + _BAND_BELOW, split_at=1000.0)
    assert result.qc.severity == Severity.PASS
    assert result.qc.value is None
    assert result.qc.message == "no visual Psat provided; estimate only"


# --- validation guards -------------------------------------------------------


def test_check_above_segment_under_two_points_raises():
    above = [(20.0, 1.0)]  # only 1 point >= split_at
    below = [(5.0, 2.0), (10.0, 2.5), (15.0, 3.0)]
    with pytest.raises(InputValidationError):
        psat_breakpoint.check(above + below, split_at=18.0)


def test_check_below_segment_under_two_points_raises():
    above = [(20.0, 1.0), (25.0, 1.5), (30.0, 2.0)]
    below = [(5.0, 2.0)]  # only 1 point < split_at
    with pytest.raises(InputValidationError):
        psat_breakpoint.check(above + below, split_at=18.0)


def test_check_degenerate_identical_pressures_in_segment_raises():
    # Above segment: both points share the SAME pressure -> ss_xx=0,
    # slope undefined (would divide by zero without the guard).
    above = [(20.0, 1.0), (20.0, 1.5)]
    below = [(5.0, 2.0), (10.0, 2.5), (15.0, 3.0)]
    with pytest.raises(InputValidationError):
        psat_breakpoint.check(above + below, split_at=18.0)


# --- integration: real CCE fixture engine output ----------------------------

CCE_WB = Path("tests/fixtures/workbooks/2_CCE_Calculation_Sheet_v5_OpenSafe_A4.xlsx")


def _load_cce_results():
    wb = openpyxl.load_workbook(CCE_WB, data_only=True)
    ws = wb["CCE Calculation"]
    stages = []
    row = 16
    while ws[f"A{row}"].value is not None:
        stages.append(
            CceStage(
                step=int(ws[f"A{row}"].value),
                p=float(ws[f"B{row}"].value),
                v_cell_cc=float(ws[f"C{row}"].value),
            )
        )
        row += 1
    inputs = CceInputs(
        t_res_f=float(ws["D6"].value),
        psat_visual=float(ws["D9"].value),
        bubble_point_step=int(ws["D10"].value),
        stages=tuple(stages),
        rho_at_psat_g_cc=float(ws["J10"].value),
        reservoir_p_psia=float(ws["D5"].value),
    )
    return calculate(inputs), inputs.psat_visual


def test_psat_breakpoint_over_real_cce_fixture_is_well_formed():
    cce_results, visual_psat = _load_cce_results()

    rv_points = [(s.p, s.rel_vol) for s in cce_results.stages]

    result = psat_breakpoint.check(
        rv_points, split_at=cce_results.psat_from_data, visual_psat=visual_psat
    )

    assert isinstance(result.qc, QCResult)
    assert result.qc.severity in (Severity.PASS, Severity.REVIEW, Severity.FAIL)
    assert math.isfinite(result.slope_above)
    assert math.isfinite(result.slope_below)
    assert math.isfinite(result.intercept_above)
    assert math.isfinite(result.intercept_below)
    assert result.qc.message

    assert math.isfinite(result.psat_estimate)

    # Observed result on this fixture run (informational only -- per the
    # task brief, NOT asserted as a severity): the CCE RV curve is
    # near-flat above Psat and steepens as pressure falls below it
    # (classic CCE shape), so a straight-line breakpoint estimate from the
    # two OLS segments is only a rough cross-check against the visual
    # reading, not expected to land tightly on it the way the dissected BD
    # tool's own (much more linear) example does. Actual run recorded here:
    #   visual Psat (fixture D9, = inputs.psat_visual):  1155.73 psig
    #   psat_from_data (split_at):                       1155.73 psig
    #   slope_above=-1.0502e-05, intercept_above=1.01031
    #   slope_below=-2.2632e-03, intercept_below=3.47064
    #   psat_estimate:                                   1092.1865 psig
    #   parallel_warning:                                False
    #   qc.value (|estimate - visual|):                  63.5435 psi
    #   qc.severity:                                     FAIL (as expected
    #     for a straight-line cross-check against a curved RV series --
    #     this is exactly why the brief asks only for well-formedness
    #     here, not a severity assertion).
    assert result.psat_estimate == pytest.approx(1092.1865060239027, rel=1e-6)
