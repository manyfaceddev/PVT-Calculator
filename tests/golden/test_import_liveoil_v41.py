"""Golden/integration tests for the LiveOil v4.1 Excel importer
(`pvt.io.excel_import.liveoil_v41`).

`test_import_then_recombine_reproduces_workbook` is transcribed verbatim from
the task-8 brief. `test_wrong_file_rejected` proves the Flash v6.1 workbook
(which also happens to carry a sheet literally named "Recombination", with a
completely different layout) is rejected rather than silently misread. The
remaining tests are supplementary: they cross-check the full cell map against
already-golden fixtures (`tests/fixtures/sa372.py`, Task 6's
`tests/golden/test_molar_recombination_sa372.py` /
`tests/golden/test_loading_sa372.py`) and cover the negative-composition
guard, the shifted-header wrong-file guard, the missing-sheet wrong-file
guard, and the unrecognized-GOR-basis guard, all needed for the
--cov-fail-under=100 branch-coverage gate.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from pvt.core.components import KATZ_FIROOZABADI as KF
from pvt.core.exceptions import InputValidationError
from pvt.experiments.recombination.loading import LoadingInputs
from pvt.experiments.recombination.molar import GorBasis, molar_split, wellstream
from pvt.io.excel_import.liveoil_v41 import read
from tests.fixtures import sa372

WB = Path("tests/fixtures/workbooks/ADRIC_LiveOil_Preparation_Calc_v4.1.xlsx")
FLASH_WB = Path("tests/fixtures/workbooks/ADRIC_Flash_Separation_Calc_v6.1.xlsx")


def test_import_then_recombine_reproduces_workbook():
    imp = read(WB)
    assert imp.gor == 339.0 and imp.z_std == 0.99
    split = molar_split(imp.gor, imp.gor_basis, imp.shrinkage, imp.sto_density_60f,
                        imp.sto_stream.mw_from_mol(), imp.gas_stream.mw_from_mol(),
                        z_std=imp.z_std)
    assert split.f_gas == pytest.approx(0.370636, abs=1e-4)
    ws = wellstream(split, imp.sto_stream, imp.gas_stream)
    assert ws.normalized_mol()["C1"] == pytest.approx(23.17, abs=0.05)


def test_wrong_file_rejected():
    with pytest.raises(InputValidationError):
        read(FLASH_WB)


def test_recombination_inputs_mapped():
    imp = read(WB)
    assert imp.gor_basis == GorBasis.SEPARATOR
    assert imp.shrinkage == pytest.approx(1.0)
    assert imp.sto_density_60f == pytest.approx(0.8196)
    assert imp.c36_mw == pytest.approx(635.0)


def test_composition_cell_map_matches_sa372_fixture():
    # Cross-check against the golden fixture already proven by Task 6's
    # test_molar_recombination_sa372.py against this same workbook.
    imp = read(WB)
    assert imp.sto_stream.mol_pct == sa372.STO_MOL_PCT
    assert imp.gas_stream.mol_pct == sa372.GAS_MOL_PCT
    assert imp.sto_stream.library.get("C36+").mw == pytest.approx(sa372.STO_C36_MW)


def test_sample_metadata_mapped():
    imp = read(WB)
    s = imp.sample
    assert s.sample_id == "SA-372"
    assert s.well == "WELL-X"
    assert s.field_name == "Upper Zakum"
    assert s.reservoir == "Kharaib-2"
    assert s.client == "ADNOC Onshore"
    assert s.cylinder == "RF1168636"
    assert s.project == "ADRIC-PVT-2026-014"
    assert s.fluid_type == "Black Oil"
    assert s.depth_ft_md == pytest.approx(9105.0)


def test_loading_inputs_mapped():
    imp = read(WB)
    # Cross-check against Task 6's golden LoadingInputs fixture (dataclass
    # equality checks all 8 fields at once).
    assert imp.loading == LoadingInputs(
        cylinder_volume_cc=1000.0, target_oil_cc=150.0,
        oil_load_p_psig=2000.0, oil_load_t_f=75.0,
        gas_load_p_psig=5000.0, gas_load_t_f=75.0, z_gas_load=0.85,
        sto_density_at_load_g_cc=0.885,
    )


def _doctored_workbook(
    tmp_path: Path,
    *,
    header_cell: tuple[str, str, str] | None = None,
    composition_override: tuple[str, int, str, float | None] | None = None,
    basis_text: str | None = None,
    omit_sheet: str | None = None,
    blank_cell: tuple[str, str] | None = None,
) -> Path:
    """Build a minimal but structurally valid 5-sheet LiveOil workbook,
    in-memory, for negative-path tests that must not touch the real fixture
    files. Mirrors the real template's layout (titles, row-14 composition
    headers, the 51-code composition block) closely enough to reach the code
    path under test.
    """
    wb = Workbook()
    wb.remove(wb.active)

    sample = wb.create_sheet("Sample_Info")
    sample["A1"] = (
        "ADRIC — LIVE OIL PREPARATION CALCULATION (v4.1, MG-0180 methodology, "
        "live formulas + physics notes)"
    )
    metadata = {
        "B5": "ADNOC Onshore", "E5": "Upper Zakum",
        "B6": "WELL-X", "E6": "Kharaib-2",
        "B7": "SA-372", "E7": "RF1168636",
        "B8": 9105, "E8": "2026-04-12",
        "B9": "Black Oil", "E9": "ADRIC-PVT-2026-014",
    }
    for addr, value in metadata.items():
        sample[addr] = value

    def _composition_sheet(name: str, title: str) -> None:
        ws = wb.create_sheet(name)
        ws["A1"] = title
        header = {
            "A14": "#", "B14": "Code", "C14": "Component", "D14": "MW (g/mol)",
            "I14": "Mol% (INPUT)",
        }
        for addr, value in header.items():
            ws[addr] = value
        codes = [code for code in KF.codes if code != "TMB124"]
        for i, code in enumerate(codes):
            row = 15 + i
            ws[f"B{row}"] = code
            ws[f"C{row}"] = code
            ws[f"D{row}"] = KF.get(code).mw
            # Row 15 (i=0) seeded non-zero (not the row any current test
            # blanks/negates) so a doctored workbook always yields a
            # non-empty CompositionStream, even when a test blanks/negates
            # one specific other row's cell.
            ws[f"I{row}"] = 50.0 if i == 0 else 0
        ws["D65"] = 635

    _composition_sheet(
        "STO_Composition",
        "STOCK TANK OIL — COMPOSITION & PROPERTIES (to C36+, Katz-Firoozabadi reference)",
    )
    _composition_sheet(
        "Gas_Composition",
        "SEPARATOR GAS — COMPOSITION & PROPERTIES (to C36+, Katz-Firoozabadi reference)",
    )
    wb["STO_Composition"]["B5"] = 0.8196

    rec = wb.create_sheet("Recombination")
    rec["A1"] = "RECOMBINATION — WELLSTREAM COMPOSITION & K-VALUE VALIDATION"
    rec["B5"] = 339.0
    rec["B6"] = basis_text if basis_text is not None else "Separator"
    rec["B7"] = 1.0
    rec["B12"] = 0.99

    load = wb.create_sheet("Loading_Volumes")
    load["A1"] = "LOADING VOLUMES — CYLINDER CHARGING FOR LIVE OIL PREPARATION"
    loading_values = {
        "B5": 1000.0, "B6": 150.0, "B7": 2000.0, "B8": 75.0,
        "B9": 5000.0, "B10": 75.0, "B11": 0.85, "B12": 0.885,
    }
    for addr, value in loading_values.items():
        load[addr] = value

    if header_cell is not None:
        sheet_name, addr, value = header_cell
        wb[sheet_name][addr] = value
    if composition_override is not None:
        sheet_name, row, col, value = composition_override
        wb[sheet_name][f"{col}{row}"] = value
    if omit_sheet is not None:
        wb.remove(wb[omit_sheet])
    if blank_cell is not None:
        sheet_name, addr = blank_cell
        wb[sheet_name][addr] = None

    path = tmp_path / "doctored.xlsx"
    wb.save(str(path))
    return path


def test_negative_composition_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, composition_override=("STO_Composition", 20, "I", -1.0))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "negative" in str(exc_info.value).lower()


def test_blank_composition_cell_is_treated_as_absent(tmp_path):
    """A blank (None) composition cell must not crash the `value < 0` sign
    check with a raw TypeError -- it is treated as absent, same as an
    explicit zero, and the component is simply missing from that stream's
    dict rather than blocking the import."""
    path = _doctored_workbook(tmp_path, composition_override=("STO_Composition", 20, "I", None))
    imp = read(path)
    missing_code = [code for code in KF.codes if code != "TMB124"][5]
    assert missing_code not in imp.sto_stream.mol_pct


def test_blank_required_cell_raises_typed_error_naming_the_cell(tmp_path):
    """A blank REQUIRED scalar cell (Recombination!B5, GOR) must raise a
    typed InputValidationError naming the exact cell address rather than
    crashing `float(None)` with a raw TypeError."""
    path = _doctored_workbook(tmp_path, blank_cell=("Recombination", "B5"))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "Recombination!B5" in str(exc_info.value)
    assert "blank or non-numeric" in str(exc_info.value)


def test_blank_sto_density_cell_raises_typed_error(tmp_path):
    """Same guard, covering read()'s own combined c36_mw/sto_density_60f
    `_num` reads (STO_Composition!B5) rather than _read_recombination_inputs's."""
    path = _doctored_workbook(tmp_path, blank_cell=("STO_Composition", "B5"))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "STO_Composition!B5" in str(exc_info.value)


def test_blank_loading_cell_raises_typed_error(tmp_path):
    """Same guard, covering _read_loading's own `_num` calls
    (Loading_Volumes!B5, cylinder_volume_cc)."""
    path = _doctored_workbook(tmp_path, blank_cell=("Loading_Volumes", "B5"))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "Loading_Volumes!B5" in str(exc_info.value)


def test_blank_sample_depth_cell_raises_typed_error(tmp_path):
    """Same guard, covering _read_sample's own `_num` call
    (Sample_Info!B8, sampling depth)."""
    path = _doctored_workbook(tmp_path, blank_cell=("Sample_Info", "B8"))
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "Sample_Info!B8" in str(exc_info.value)


def test_unexpected_header_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, header_cell=("STO_Composition", "B14", "WRONG"))
    with pytest.raises(InputValidationError):
        read(path)


def test_missing_sheet_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, omit_sheet="Loading_Volumes")
    with pytest.raises(InputValidationError):
        read(path)


def test_unrecognized_gor_basis_rejected(tmp_path):
    path = _doctored_workbook(tmp_path, basis_text="Sideways")
    with pytest.raises(InputValidationError) as exc_info:
        read(path)
    assert "basis" in str(exc_info.value).lower()
